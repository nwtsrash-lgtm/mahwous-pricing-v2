"""
محرك كشط المنافسين عبر mahally.com — سريع ودقيق واحترافي
=======================================================
يستخدم بيانات Algolia المضمنة في صفحات mahally.com لاستخراج
بيانات المنتجات (الأسعار، العلامات، التصنيفات) من المتاجر المنافسة.

الطريقة المُثبتة:
  1. جلب صفحة المتجر: https://mahally.com/stores/{store_id}?page=N
  2. استخراج JSON من window[Symbol.for('InstantSearchInitialResults')]
  3. تحليل hits من products_v2_store_view → results[0]
"""
import sys, io
if not isinstance(sys.stdout, io.TextIOWrapper) or sys.stdout.encoding != 'utf-8':
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
    except Exception:
        pass


import requests, re, json, time, logging, unicodedata, sqlite3, csv, os, threading
from datetime import datetime
from typing import List, Dict, Optional, Callable
from concurrent.futures import ThreadPoolExecutor, as_completed

log = logging.getLogger(__name__)

# ── ثوابت ──────────────────────────────────────────────────────────────
BASE_URL = "https://mahally.com/stores/{store_id}?page={page}"
ALGOLIA_INDEX = "products_v2_store_view"
MAX_PAGES = 10
HITS_PER_PAGE = 1000
PAGE_DELAY = 0.3

# ── Algolia API مباشرة (أسرع 100x وبدون حظر) ──────────────────────────
ALGOLIA_APP_ID = os.environ.get("ALGOLIA_APP_ID", "L41Y35UONW")
ALGOLIA_API_KEY = os.environ.get("ALGOLIA_API_KEY", "f60e98a284e4b402af626d0dd1fc6cbd")
ALGOLIA_API_URL = f"https://{ALGOLIA_APP_ID}-dsn.algolia.net/1/indexes/{ALGOLIA_INDEX}/query"
ALGOLIA_HEADERS = {
    "X-Algolia-Application-Id": ALGOLIA_APP_ID,
    "X-Algolia-API-Key": ALGOLIA_API_KEY,
    "Content-Type": "application/json",
}

CHROME_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/131.0.0.0 Safari/537.36"
)

DEFAULT_HEADERS = {
    "User-Agent": CHROME_UA,
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "ar,en-US;q=0.9,en;q=0.8",
    "Accept-Encoding": "gzip, deflate, br",
    "Cache-Control": "no-cache",
}

# ── المتاجر المعروفة ──────────────────────────────────────────────────
KNOWN_STORES = {
    216339537:    "سعيد صلاح",
    1951545756:   "سارا ميك اب",
    1891860617:   "عالم جيفينشي",
    997023036:    "الفاخرة للنيش",
    1313514692:   "فانيلا",
}

# ── مسار ملف المنافسين ─────────────────────────────────────────────────
_DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")
COMPETITORS_JSON = os.path.join(_DATA_DIR, "competitors_list_v30.json")


class MahallyScraper:
    """
    محرك كشط احترافي لاستخراج بيانات المنتجات من متاجر mahally.com.

    الاستخدام:
        scraper = MahallyScraper(db_path="data/pricing.db")
        products = scraper.scrape_store(216339537, "سعيد صلاح")
        all_data = scraper.scrape_all_stores()
        scraper.export_csv(all_data, "exports/")
        scraper.export_excel(all_data, "exports/")
    """

    def __init__(self, db_path: str = None, progress_callback: Callable = None):
        """
        تهيئة محرك الكشط.

        Args:
            db_path: مسار قاعدة البيانات (اختياري)
            progress_callback: دالة إشعار التقدم بالشكل:
                callback(store_name, current_page, total_pages, message)
        """
        # ── قفل حماية seen_ids من سباق الخيوط ──
        self._seen_lock = threading.Lock()
        # ── الجلسة ──
        self.session = requests.Session()
        self.session.headers.update(DEFAULT_HEADERS)

        # ── قاعدة البيانات ──
        self.db_path = db_path

        # ── دالة التقدم ──
        self.progress_callback = progress_callback

        # ── تحميل قائمة المتاجر ──
        self.stores: Dict[int, str] = dict(KNOWN_STORES)
        self._load_competitors_from_json()

        log.info(
            "MahallyScraper initialized — %d stores configured", len(self.stores)
        )

    # ──────────────────────────────────────────────────────────────────────
    #  API العام
    # ──────────────────────────────────────────────────────────────────────

    def scrape_store(self, store_id: int, store_name: str = None) -> List[Dict]:
        """
        كشط متجر واحد — يستخدم Algolia API مباشرة (الأسرع).
        إذا فشل، يعود للطريقة القديمة (HTML scraping).
        """
        store_name = store_name or self.stores.get(store_id, str(store_id))
        log.info("▶ بدء كشط '%s' (ID: %d)", store_name, store_id)

        # ── الطريقة 1: Algolia API مباشرة (أسرع 100x) ──
        products = self._scrape_via_algolia(store_id, store_name)
        if products:
            log.info("✔ '%s' — %d منتج (Algolia API)", store_name, len(products))
            return products

        # ── الطريقة 2: HTML scraping (fallback) ──
        log.info("  ⚠ Algolia فشل — محاولة HTML scraping")
        products = self._scrape_via_html(store_id, store_name)
        log.info("✔ '%s' — %d منتج (HTML)", store_name, len(products))
        return products

    def _scrape_via_algolia(self, store_id: int, store_name: str) -> List[Dict]:
        """كشط عبر Algolia API — تقسيم بنطاقات الأسعار لجلب 100% من المنتجات."""
        seen_ids: set = set()
        products: List[Dict] = []

        # أولاً: عدد المنتجات المتوقع
        try:
            r0 = requests.post(
                ALGOLIA_API_URL, headers=ALGOLIA_HEADERS,
                json={"query": "", "hitsPerPage": 0,
                      "facetFilters": [f"store_id:{store_id}"]},
                timeout=15,
            )
            expected = r0.json().get("nbHits", 0) if r0.status_code == 200 else 0
        except Exception:
            expected = 0

        if expected == 0:
            return []

        log.info("  Algolia: %d منتج متوقع", expected)

        # نطاقات سعرية ذكية
        ranges = []
        for s in range(0, 500, 50):
            ranges.append((s, s + 50))
        for s in range(500, 2000, 200):
            ranges.append((s, s + 200))
        for s in range(2000, 10000, 1000):
            ranges.append((s, s + 1000))
        ranges.append((10000, 999999))

        # تنفيذ متوازي — 5 نطاقات في وقت واحد
        from concurrent.futures import ThreadPoolExecutor, as_completed

        def _fetch_range(lo, hi):
            return self._algolia_range(store_id, store_name, lo, hi, seen_ids)

        with ThreadPoolExecutor(max_workers=5) as ex:
            futures = {ex.submit(_fetch_range, lo, hi): (lo, hi) for lo, hi in ranges}
            for f in as_completed(futures):
                try:
                    prods = f.result()
                    products.extend(prods)
                except Exception as e:
                    log.warning("  range error: %s", e)

        coverage = (len(products) / expected * 100) if expected else 0
        self._notify(store_name, 1, 1, f"{len(products):,}/{expected:,} ({coverage:.0f}%)")
        log.info("  %d/%d (%.1f%%)", len(products), expected, coverage)
        return products

    def _algolia_range(self, store_id, store_name, lo, hi, seen_ids):
        """جلب منتجات نطاق سعري — تقسيم تلقائي إذا > 1000."""
        filt = f"store_id:{store_id} AND price.SA.SAR:{lo} TO {hi}"
        products = []

        for attempt in range(3):
            try:
                resp = requests.post(
                    ALGOLIA_API_URL, headers=ALGOLIA_HEADERS,
                    json={
                        "query": "", "hitsPerPage": HITS_PER_PAGE, "page": 0,
                        "filters": filt,
                        "attributesToRetrieve": [
                            "name", "price", "regular_price", "sale_price",
                            "brand_name", "image", "categories", "discount_percentage",
                            "public_product_id", "store_id", "purchasable",
                        ],
                    },
                    timeout=20,
                )
                if resp.status_code == 200:
                    data = resp.json()
                    hits = data.get("hits", [])
                    nb = data.get("nbHits", 0)

                    for h in hits:
                        pid = h.get("public_product_id", h.get("objectID", ""))
                        with self._seen_lock:
                            if pid and pid not in seen_ids:
                                seen_ids.add(pid)
                                is_new = True
                            else:
                                is_new = False
                        if is_new:
                            prod = self._parse_hit(h, store_name)
                            if prod:
                                products.append(prod)

                    # تقسيم تلقائي إذا النطاق > 1000
                    if nb > 1000 and (hi - lo) > 2:
                        mid = (lo + hi) / 2
                        products += self._algolia_range(store_id, store_name, lo, mid, seen_ids)
                        products += self._algolia_range(store_id, store_name, mid, hi, seen_ids)
                    break
                else:
                    time.sleep(1)
            except Exception:
                time.sleep(1)

        return products

    def _scrape_via_html(self, store_id: int, store_name: str) -> List[Dict]:
        """كشط عبر HTML من mahally.com (fallback)."""
        products: List[Dict] = []
        page = 1
        total_pages = MAX_PAGES

        while page <= min(total_pages, MAX_PAGES):
            url = BASE_URL.format(store_id=store_id, page=page)
            success = False

            for attempt in range(3):
                try:
                    resp = self.session.get(url, timeout=30)
                    if resp.status_code == 503:
                        wait = [3, 8, 15][attempt]
                        log.warning("  HTML page %d — 503 (attempt %d) — wait %ds", page, attempt+1, wait)
                        time.sleep(wait)
                        continue
                    if resp.status_code != 200:
                        break

                    hits, nb_hits, nb_pages = self._extract_hits(resp.text)
                    if page == 1:
                        total_pages = min(nb_pages, MAX_PAGES)
                    if not hits:
                        break

                    for hit in hits:
                        prod = self._parse_hit(hit, store_name)
                        if prod:
                            products.append(prod)
                    success = True
                    break
                except Exception as e:
                    log.error("  HTML page %d error: %s", page, e)
                    time.sleep(3)

            if not success and page > 1:
                break
            page += 1
            if page <= total_pages:
                time.sleep(max(PAGE_DELAY, 1.0))

        return products

    def scrape_all_stores(
        self, store_ids: Dict[int, str] = None
    ) -> Dict[str, List[Dict]]:
        """
        كشط جميع المتاجر المُسجلة بالتوازي.

        Args:
            store_ids: قاموس {store_id: store_name} — إذا لم يُحدد تُستخدم القائمة المعروفة

        Returns:
            قاموس {store_name: [products]}
        """
        targets = store_ids or self.stores
        results: Dict[str, List[Dict]] = {}
        total_stores = len(targets)

        log.info("🚀 بدء كشط %d متجر بالتوازي (max_workers=3)", total_stores)

        with ThreadPoolExecutor(max_workers=3) as executor:
            futures = {
                executor.submit(self.scrape_store, sid, sname): (sid, sname)
                for sid, sname in targets.items()
            }
            for future in as_completed(futures):
                sid, sname = futures[future]
                try:
                    products = future.result()
                    results[sname] = products
                    log.info(
                        "  ✅ %s — %d منتج", sname, len(products)
                    )
                except Exception as e:
                    log.error("  ❌ %s — خطأ: %s", sname, e)
                    results[sname] = []

        total_products = sum(len(v) for v in results.values())
        log.info(
            "🏁 انتهى الكشط — %d متجر، %d منتج إجمالي",
            len(results), total_products,
        )
        return results

    def save_to_db(self, products: List[Dict], competitor_name: str) -> int:
        """
        حفظ المنتجات في جدول competitor_products_store.

        Args:
            products: قائمة قواميس المنتجات
            competitor_name: اسم المنافس

        Returns:
            عدد السجلات المحفوظة/المحدثة
        """
        if not products or not self.db_path:
            return 0

        conn = sqlite3.connect(self.db_path)
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS competitor_products_store (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                competitor TEXT NOT NULL,
                product_name TEXT NOT NULL,
                norm_name TEXT NOT NULL,
                price REAL DEFAULT 0,
                image_url TEXT DEFAULT '',
                product_url TEXT DEFAULT '',
                brand TEXT DEFAULT '',
                size TEXT DEFAULT '',
                gender TEXT DEFAULT '',
                added_at TEXT DEFAULT (datetime('now','localtime')),
                updated_at TEXT DEFAULT (datetime('now','localtime')),
                first_seen_at TEXT DEFAULT (datetime('now','localtime')),
                rating_count INTEGER DEFAULT 0,
                discount_pct REAL DEFAULT 0,
                original_price REAL DEFAULT 0,
                category TEXT DEFAULT '',
                sku TEXT DEFAULT '',
                UNIQUE(competitor, norm_name)
            )
        """)
        # إضافة أعمدة جديدة إذا لم تكن موجودة (ترقية DB)
        for col_def in [
            ("first_seen_at", "TEXT DEFAULT (datetime('now','localtime'))"),
            ("rating_count", "INTEGER DEFAULT 0"),
            ("discount_pct", "REAL DEFAULT 0"),
            ("original_price", "REAL DEFAULT 0"),
            ("category", "TEXT DEFAULT ''"),
            ("sku", "TEXT DEFAULT ''"),
        ]:
            # whitelist: اسم العمود معرّف صالح + النوع الأساسي ضمن المسموح
            import re as _re
            if not _re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", col_def[0]):
                continue
            if col_def[1].split()[0].upper() not in {"TEXT", "INTEGER", "REAL", "BLOB", "NUMERIC"}:
                continue
            try:
                cur.execute(f"ALTER TABLE competitor_products_store ADD COLUMN {col_def[0]} {col_def[1]}")
            except sqlite3.OperationalError:
                pass  # العمود موجود مسبقاً

        # إضافة فهارس لتسريع البحث
        # L2/C3: أُزيل خلق idx_cps_competitor/brand/price — مكرّرة (يملكها
        # db_manager.ensure_indexes عبر idx_cps_*2 + المركّب) فكانت تُعيد التكرار
        # بعد إسقاط C1. نُبقي first_seen/rating (الفهرسان القانونيان لعموديهما).
        for idx_sql in [
            "CREATE INDEX IF NOT EXISTS idx_cps_first_seen ON competitor_products_store(first_seen_at)",
            "CREATE INDEX IF NOT EXISTS idx_cps_rating ON competitor_products_store(rating_count)",
        ]:
            try:
                cur.execute(idx_sql)
            except sqlite3.OperationalError:
                pass

        saved = 0
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for p in products:
            name = p.get("name", "")
            if len(name) < 2:
                continue
            norm = self.normalize(name)
            try:
                existing = cur.execute(
                    "SELECT id FROM competitor_products_store "
                    "WHERE competitor=? AND norm_name=?",
                    (competitor_name, norm),
                ).fetchone()

                if existing:
                    row_id = existing[0]
                    cur.execute(
                        """UPDATE competitor_products_store
                           SET price=?, updated_at=?,
                               image_url=COALESCE(NULLIF(?,''),
                                   (SELECT image_url FROM competitor_products_store WHERE id=?)),
                               product_url=COALESCE(NULLIF(?,''),
                                   (SELECT product_url FROM competitor_products_store WHERE id=?)),
                               brand=COALESCE(NULLIF(?,''),
                                   (SELECT brand FROM competitor_products_store WHERE id=?)),
                               rating_count=?, discount_pct=?, original_price=?,
                               category=COALESCE(NULLIF(?,''),
                                   (SELECT category FROM competitor_products_store WHERE id=?)),
                               sku=COALESCE(NULLIF(?,''),
                                   (SELECT sku FROM competitor_products_store WHERE id=?))
                           WHERE id=?""",
                        (
                            p["price"], now,
                            p.get("image", ""), row_id,
                            p.get("url", ""), row_id,
                            p.get("brand", ""), row_id,
                            p.get("rating_count", 0),
                            p.get("discount_pct", 0),
                            p.get("original_price", 0),
                            p.get("category", ""), row_id,
                            p.get("sku", ""), row_id,
                            row_id,
                        ),
                    )
                else:
                    cur.execute(
                        """INSERT INTO competitor_products_store
                           (competitor, product_name, norm_name, price,
                            image_url, product_url, brand, size, gender,
                            added_at, updated_at, first_seen_at,
                            rating_count, discount_pct, original_price,
                            category, sku)
                           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (
                            competitor_name, name, norm, p["price"],
                            p.get("image", ""), p.get("url", ""),
                            p.get("brand", ""), "", "", now, now, now,
                            p.get("rating_count", 0),
                            p.get("discount_pct", 0),
                            p.get("original_price", 0),
                            p.get("category", ""),
                            p.get("sku", ""),
                        ),
                    )
                saved += 1
            except sqlite3.Error as e:
                log.debug("DB error for '%s': %s", name[:30], e)

        conn.commit()
        conn.close()
        log.info(
            "💾 حُفظ %d/%d منتج لـ '%s'", saved, len(products), competitor_name
        )
        return saved

    def export_csv(self, data: Dict[str, List[Dict]], export_dir: str) -> str:
        """
        تصدير البيانات بصيغة CSV (نفس تنسيق competitors_latest.csv).

        Args:
            data: قاموس {store_name: [products]}
            export_dir: مجلد التصدير

        Returns:
            مسار الملف المُدمج
        """
        os.makedirs(export_dir, exist_ok=True)
        now = datetime.now().isoformat()
        headers = [
            "store", "name", "price", "original_price", "sku",
            "url", "image", "brand", "category", "availability", "scraped_at",
        ]

        # ── ملف مُدمج ──
        combined_path = os.path.join(export_dir, "competitors_latest.csv")
        with open(combined_path, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(headers)
            for store_name, products in data.items():
                for p in products:
                    w.writerow([
                        store_name,
                        p["name"],
                        p["price"],
                        p["original_price"],
                        p.get("sku", ""),
                        p.get("url", ""),
                        p.get("image", ""),
                        p.get("brand", ""),
                        p.get("category", ""),
                        str(p.get("availability", True)).lower(),
                        now,
                    ])
        log.info("📄 CSV مُدمج: %s", combined_path)

        # ── ملفات لكل متجر ──
        per_dir = os.path.join(export_dir, "per_store")
        os.makedirs(per_dir, exist_ok=True)
        for store_name, products in data.items():
            if not products:
                continue
            safe_name = re.sub(r"[^\w\-]", "_", store_name)
            path = os.path.join(per_dir, f"{safe_name}.csv")
            with open(path, "w", encoding="utf-8-sig", newline="") as f:
                w = csv.writer(f)
                w.writerow(headers)
                for p in products:
                    w.writerow([
                        store_name,
                        p["name"],
                        p["price"],
                        p["original_price"],
                        p.get("sku", ""),
                        p.get("url", ""),
                        p.get("image", ""),
                        p.get("brand", ""),
                        p.get("category", ""),
                        str(p.get("availability", True)).lower(),
                        now,
                    ])
            log.info("  📄 %s: %d منتج → %s", store_name, len(products), path)

        return combined_path

    def export_excel(self, data: Dict[str, List[Dict]], export_dir: str) -> Optional[str]:
        """
        تصدير احترافي بصيغة Excel مع دعم RTL وتنسيق مميز.

        Args:
            data: قاموس {store_name: [products]}
            export_dir: مجلد التصدير

        Returns:
            مسار ملف Excel أو None عند الفشل
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            log.warning("⚠ مكتبة openpyxl غير متوفرة — تخطي تصدير Excel")
            return None

        os.makedirs(export_dir, exist_ok=True)
        wb = openpyxl.Workbook()

        # ── أنماط ──
        header_fill = PatternFill(
            start_color="1F4E79", end_color="1F4E79", fill_type="solid"
        )
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        def _style_header(ws, headers_list):
            """تنسيق صف الرأس"""
            for col_idx, header_text in enumerate(headers_list, 1):
                cell = ws.cell(row=1, column=col_idx, value=header_text)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
                cell.border = thin_border

        # ── ورقة الملخص ──
        ws_summary = wb.active
        ws_summary.title = "ملخص"
        ws_summary.sheet_view.rightToLeft = True
        summary_headers = [
            "#", "المتجر", "المنتجات", "أقل سعر",
            "أعلى سعر", "متوسط السعر", "نسبة الخصم",
        ]
        _style_header(ws_summary, summary_headers)

        row = 2
        for idx, (store_name, products) in enumerate(data.items(), 1):
            prices = [p["price"] for p in products if p.get("price", 0) > 0]
            discounts = [
                p.get("discount", 0) for p in products
                if p.get("discount", 0) > 0
            ]
            ws_summary.cell(row=row, column=1, value=idx)
            ws_summary.cell(row=row, column=2, value=store_name)
            ws_summary.cell(row=row, column=3, value=len(products))
            ws_summary.cell(
                row=row, column=4,
                value=round(min(prices), 2) if prices else 0,
            )
            ws_summary.cell(
                row=row, column=5,
                value=round(max(prices), 2) if prices else 0,
            )
            ws_summary.cell(
                row=row, column=6,
                value=round(sum(prices) / len(prices), 2) if prices else 0,
            )
            ws_summary.cell(
                row=row, column=7,
                value=(
                    f"{round(sum(discounts) / len(discounts))}%"
                    if discounts else "0%"
                ),
            )
            row += 1

        # عرض الأعمدة
        summary_widths = {"A": 5, "B": 25, "C": 12, "D": 12, "E": 12, "F": 14, "G": 12}
        for col_letter, width in summary_widths.items():
            ws_summary.column_dimensions[col_letter].width = width

        # ── ورقة لكل متجر ──
        store_headers = [
            "#", "اسم المنتج", "السعر", "السعر الأصلي",
            "الماركة", "التصنيف", "الخصم%", "SKU", "الرابط",
        ]
        store_widths = {"A": 5, "B": 50, "C": 10, "D": 12, "E": 20, "F": 25, "G": 8, "H": 15, "I": 40}

        for store_name, products in data.items():
            if not products:
                continue
            safe_title = re.sub(r"[^\w]", "_", store_name)[:31]  # Excel max 31 chars
            ws = wb.create_sheet(title=safe_title)
            ws.sheet_view.rightToLeft = True
            _style_header(ws, store_headers)

            for i, p in enumerate(products, 1):
                ws.cell(row=i + 1, column=1, value=i)
                ws.cell(row=i + 1, column=2, value=p["name"])
                ws.cell(row=i + 1, column=3, value=p["price"])
                ws.cell(row=i + 1, column=4, value=p["original_price"])
                ws.cell(row=i + 1, column=5, value=p.get("brand", ""))
                ws.cell(row=i + 1, column=6, value=p.get("category", ""))
                ws.cell(row=i + 1, column=7, value=f"{p.get('discount', 0)}%")
                ws.cell(row=i + 1, column=8, value=p.get("sku", ""))
                ws.cell(row=i + 1, column=9, value=p.get("url", ""))

            for col_letter, width in store_widths.items():
                ws.column_dimensions[col_letter].width = width

        xlsx_path = os.path.join(export_dir, "competitors_all_products.xlsx")
        wb.save(xlsx_path)
        log.info("📊 Excel: %s", xlsx_path)
        return xlsx_path

    def get_store_info(self, store_id: int) -> Dict:
        """
        جلب سريع لمعلومات المتجر من الصفحة الأولى.

        Args:
            store_id: معرف المتجر

        Returns:
            قاموس {name, store_id, total_products, total_pages}
        """
        url = BASE_URL.format(store_id=store_id, page=1)
        try:
            resp = self.session.get(url, timeout=30)
            if resp.status_code != 200:
                return {"store_id": store_id, "error": f"HTTP {resp.status_code}"}

            hits, nb_hits, nb_pages = self._extract_hits(resp.text)

            # محاولة استخراج اسم المتجر من العنوان
            title_match = re.search(r"<title>([^<]+)</title>", resp.text)
            name = title_match.group(1).strip() if title_match else str(store_id)

            return {
                "name": name,
                "store_id": store_id,
                "total_products": nb_hits,
                "total_pages": nb_pages,
                "sample_count": len(hits),
            }
        except Exception as e:
            return {"store_id": store_id, "error": str(e)}

    # ──────────────────────────────────────────────────────────────────────
    #  الدوال الداخلية
    # ──────────────────────────────────────────────────────────────────────

    def _extract_hits(self, html: str) -> tuple:
        """
        استخراج (hits, nbHits, nbPages) من HTML صفحة المتجر.

        يبحث عن window[Symbol.for('InstantSearchInitialResults')] ثم
        يستخرج JSON المضمّن بين = { ... };

        Returns:
            (hits: list, nbHits: int, nbPages: int)
        """
        hits, nb_hits, nb_pages = [], 0, 0

        # ── الطريقة الأولى: استخراج JSON الكامل ──
        # البحث عن بداية الكائن بعد InstantSearchInitialResults
        marker = 'InstantSearchInitialResults'
        marker_pos = html.find(marker)
        if marker_pos == -1:
            log.warning("  لم يُعثر على InstantSearchInitialResults")
            return hits, nb_hits, nb_pages

        # البحث عن = { بعد العلامة
        eq_pos = html.find("=", marker_pos + len(marker))
        if eq_pos == -1:
            return hits, nb_hits, nb_pages

        brace_pos = html.find("{", eq_pos)
        if brace_pos == -1:
            return hits, nb_hits, nb_pages

        # إيجاد القوس المطابق }
        json_str = self._find_matching_brace(html, brace_pos)
        if not json_str:
            # طريقة بديلة: regex
            match = re.search(
                r'InstantSearchInitialResults["\'\]]+\s*=\s*(\{.*?\});\s*</script>',
                html, re.S,
            )
            if match:
                json_str = match.group(1)
            else:
                return hits, nb_hits, nb_pages

        try:
            data = json.loads(json_str)
        except json.JSONDecodeError:
            log.warning("  فشل تحليل JSON (%d حرف)", len(json_str))
            return hits, nb_hits, nb_pages

        # ── التنقل في بنية البيانات ──
        results = data.get(ALGOLIA_INDEX, {}).get("results", [])
        if not results:
            # محاولة أي مفتاح يحتوي results
            for key in data:
                if isinstance(data[key], dict) and "results" in data[key]:
                    results = data[key]["results"]
                    break

        if not results:
            return hits, nb_hits, nb_pages

        result = results[0] if isinstance(results, list) else results
        hits = result.get("hits", [])
        nb_hits = result.get("nbHits", 0)
        nb_pages = result.get("nbPages", 0)

        return hits, nb_hits, nb_pages

    @staticmethod
    def _find_matching_brace(text: str, start: int) -> Optional[str]:
        """
        إيجاد النص بين { المطابق } بدءاً من الموضع start.
        يتعامل مع الأقواس المتداخلة والنصوص بين علامات الاقتباس.
        """
        if start >= len(text) or text[start] != "{":
            return None

        depth = 0
        in_string = False
        escape = False
        i = start

        while i < len(text):
            ch = text[i]

            if escape:
                escape = False
                i += 1
                continue

            if ch == "\\":
                escape = True
                i += 1
                continue

            if ch == '"' and not escape:
                in_string = not in_string
                i += 1
                continue

            if in_string:
                i += 1
                continue

            if ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    return text[start : i + 1]

            i += 1

        return None

    def _parse_hit(self, hit: dict, store_name: str) -> Optional[Dict]:
        """
        تحويل عنصر Algolia واحد إلى قاموس منتج مُوحد.

        Args:
            hit: كائن المنتج من Algolia
            store_name: اسم المتجر

        Returns:
            قاموس المنتج أو None إذا كان غير صالح
        """
        try:
            # ── الاسم ──
            name = ""
            name_obj = hit.get("name", {})
            if isinstance(name_obj, dict):
                name = name_obj.get("ar", name_obj.get("en", ""))
            elif isinstance(name_obj, str):
                name = name_obj
            if not name:
                name = hit.get("name_ar", "")
            if not name or len(name) < 2:
                return None

            # ── السعر الحالي ──
            price = self._extract_price(hit, "price")
            if not price or price <= 0:
                return None

            # ── السعر الأصلي ──
            regular_price = self._extract_price(hit, "regular_price") or price

            # ── سعر التخفيض ──
            sale_price = self._extract_price(hit, "sale_price") or price

            # ── الماركة ──
            brand = ""
            brand_obj = hit.get("brand_name", {})
            if isinstance(brand_obj, dict):
                brand = brand_obj.get("ar", brand_obj.get("en", ""))
            elif isinstance(brand_obj, str):
                brand = brand_obj

            # ── الصورة ──
            image = hit.get("image", "")

            # ── التصنيف ──
            category = ""
            cat_obj = hit.get("categories", {})
            if isinstance(cat_obj, dict):
                # lvl2 أكثر تفصيلاً — نفضله
                cat_val = cat_obj.get("lvl2", cat_obj.get("lvl1", cat_obj.get("lvl0", "")))
                if isinstance(cat_val, list):
                    category = cat_val[0] if cat_val else ""
                else:
                    category = str(cat_val)

            # ── معرف المنتج ──
            sku = str(hit.get("public_product_id", hit.get("objectID", "")))

            # ── رابط المنتج ──
            hit_store_id = hit.get("store_id", "")
            url = (
                f"https://mahally.com/products/{hit_store_id}/{sku}"
                if hit_store_id and sku
                else ""
            )

            # ── الخصم ──
            discount = hit.get("discount_percentage", 0) or 0

            # ── التوفر ──
            availability = hit.get("purchasable", True)

            # ── الوصف ──
            description = ""
            desc_obj = hit.get("description", {})
            if isinstance(desc_obj, dict):
                description = desc_obj.get("ar", desc_obj.get("en", ""))
            elif isinstance(desc_obj, str):
                description = desc_obj

            return {
                "store": store_name,
                "name": name.strip()[:300],
                "price": round(sale_price or price, 2),
                "original_price": round(regular_price or price, 2),
                "brand": brand.strip(),
                "image": image,
                "url": url,
                "category": category,
                "discount": round(float(discount), 2),
                "availability": availability,
                "sku": sku,
                "description": description[:500] if description else "",
                "rating_count": int(hit.get("rating_count", 0) or 0),
                "discount_pct": round(float(discount), 2),
            }
        except Exception as e:
            log.debug("خطأ في تحليل hit: %s", e)
            return None

    @staticmethod
    def _extract_price(hit: dict, field: str) -> float:
        """استخراج السعر من بنية price.SA.SAR المتداخلة."""
        price_obj = hit.get(field, {})
        if isinstance(price_obj, dict):
            sa = price_obj.get("SA", {})
            if isinstance(sa, dict):
                return float(sa.get("SAR", 0) or 0)
            return 0.0
        if isinstance(price_obj, (int, float)):
            return float(price_obj)
        if isinstance(price_obj, str):
            try:
                return float(price_obj)
            except ValueError:
                return 0.0
        return 0.0

    @staticmethod
    def normalize(text: str) -> str:
        """
        تطبيع النص العربي للمقارنة:
        - أإآا → ا
        - ةه → ه
        - يى → ي
        - إزالة التشكيل
        - توحيد المسافات
        """
        t = unicodedata.normalize("NFKC", str(text or ""))
        # إزالة التشكيل (الفتحة، الكسرة، الضمة، السكون، الشدة، التنوين)
        t = re.sub(r"[\u064B-\u065F\u0670]", "", t)
        # توحيد الألف
        t = re.sub(r"[أإآا]", "ا", t)
        # توحيد التاء والهاء
        t = re.sub(r"[ةه]", "ه", t)
        # توحيد الياء
        t = re.sub(r"[يى]", "ي", t)
        # توحيد المسافات
        return re.sub(r"\s+", " ", t).strip().lower()

    def _load_competitors_from_json(self):
        """تحميل معرفات المتاجر من ملف المنافسين إن وُجد."""
        if not os.path.exists(COMPETITORS_JSON):
            return
        try:
            with open(COMPETITORS_JSON, "r", encoding="utf-8") as f:
                competitors = json.load(f)
            for comp in competitors:
                mid = comp.get("mahally_store_id")
                if mid:
                    name = comp.get("name", str(mid))
                    self.stores[int(mid)] = name
            log.debug(
                "تم تحميل %d متجر من %s", len(self.stores), COMPETITORS_JSON
            )
        except Exception as e:
            log.debug("تعذر تحميل المنافسين: %s", e)

    def _notify(self, store_name: str, current: int, total: int, message: str):
        """استدعاء دالة التقدم إن وُجدت."""
        if self.progress_callback:
            try:
                self.progress_callback(store_name, current, total, message)
            except Exception:
                pass


# ══════════════════════════════════════════════════════════════════════════
#  تشغيل مباشر (CLI)
# ══════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)-7s | %(message)s",
        datefmt="%H:%M:%S",
    )

    print("=" * 70)
    print("   MAHALLY SCRAPER ENGINE — CLI MODE")
    print(f"   {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)

    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    db_path = os.path.join(base_dir, "data", "pricing_v18.db")
    export_dir = os.path.join(base_dir, "exports")

    def cli_progress(store, current, total, msg):
        print(f"  [{current}/{total}] {store}: {msg}")

    scraper = MahallyScraper(db_path=db_path, progress_callback=cli_progress)

    # كشط جميع المتاجر
    all_data = scraper.scrape_all_stores()

    # حفظ في قاعدة البيانات
    for store_name, products in all_data.items():
        if products:
            scraper.save_to_db(products, store_name)

    # تصدير
    scraper.export_csv(all_data, export_dir)
    scraper.export_excel(all_data, export_dir)

    # ملخص
    total = sum(len(v) for v in all_data.values())
    print(f"\n{'=' * 70}")
    print(f"  الإجمالي: {total} منتج من {len(all_data)} متجر")
    for name, prods in all_data.items():
        status = f"✅ {len(prods)}" if prods else "❌ 0"
        print(f"    {name:20s}: {status}")
    print("=" * 70)
