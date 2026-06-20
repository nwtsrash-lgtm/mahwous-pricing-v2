"""
engines/engine.py  v26.0 — محرك المطابقة الفائق السرعة
═══════════════════════════════════════════════════════
🚀 تطبيع مسبق (Pre-normalize) → vectorized cdist → Gemini للغموض فقط
⚡ 5x أسرع من v20 مع نفس الدقة 99.5%
🔧 v26.0: مرادفات موسعة + تصحيح إملائي ذكي للماركات

الخطة:
  1. عند رفع الملف → تطبيع كل منتجات المنافس مرة واحدة (cache)
  2. لكل منتجنا → cdist vectorized دفعة واحدة (بدل loop)
  3. أفضل 5 مرشحين → Gemini فقط إذا score بين 62-96%
  4. score ≥97% → تلقائي فوري  |  score <62% → مفقود
"""
import re, io, json, os, hashlib, sqlite3, gc, threading
import functools as _functools
from datetime import datetime
import pandas as pd
from utils.data_helpers import first_image_url_string
from utils.data_paths import get_data_db_path
from utils.helpers import favicon_url_for_site, fetch_og_image_url
from rapidfuzz import fuzz, process as rf_process
from rapidfuzz.distance import Indel
import requests as _req

from engines.mahwous_core import apply_strict_pipeline_filters

# ─── استيراد الإعدادات ───────────────────────
try:
    from config import (REJECT_KEYWORDS, KNOWN_BRANDS, WORD_REPLACEMENTS,
                        MATCH_THRESHOLD, HIGH_CONFIDENCE,
                        PRICE_TOLERANCE, TESTER_KEYWORDS, SET_KEYWORDS,
                        GEMINI_API_KEYS, OPENROUTER_API_KEY)
except Exception:
    REJECT_KEYWORDS = ["sample","عينة","عينه","decant","تقسيم","split","miniature"]
    KNOWN_BRANDS = [
        "Dior","Chanel","Gucci","Tom Ford","Versace","Armani","YSL","Prada","Burberry",
        "Hermes","Creed","Montblanc","Amouage","Rasasi","Lattafa","Arabian Oud","Ajmal",
        "Al Haramain","Afnan","Armaf","Mancera","Montale","Kilian","Jo Malone",
        "Carolina Herrera","Paco Rabanne","Mugler","Ralph Lauren","Parfums de Marly",
        "Nishane","Xerjoff","Byredo","Le Labo","Roja","Narciso Rodriguez",
        "Dolce & Gabbana","Valentino","Bvlgari","Cartier","Hugo Boss","Calvin Klein",
        "Givenchy","Lancome","Guerlain","Jean Paul Gaultier","Issey Miyake","Davidoff",
        "Coach","Michael Kors","Initio","Memo Paris","Maison Margiela","Diptyque",
        "Missoni","Juicy Couture","Moschino","Dunhill","Bentley","Jaguar",
        "Boucheron","Chopard","Elie Saab","Escada","Ferragamo","Fendi",
        "Kenzo","Lacoste","Loewe","Rochas","Roberto Cavalli","Tiffany",
        "Van Cleef","Azzaro","Chloe","Elizabeth Arden","Swiss Arabian",
        "Penhaligons","Clive Christian","Floris","Acqua di Parma",
        "Ard Al Zaafaran","Nabeel","Asdaaf","Maison Alhambra",
        "Tiziana Terenzi","Maison Francis Kurkdjian","Serge Lutens",
        "Frederic Malle","Ormonde Jayne","Zoologist","Tauer",
        "Banana Republic","Benetton","Bottega Veneta","Celine","Dsquared2",
        "Ermenegildo Zegna","Sisley","Mexx","Amadou","Thameen",
        "Nasomatto","Nicolai","Replica","Atelier Cologne","Aerin",
        "Angel Schlesser","Annick Goutal","Antonio Banderas","Balenciaga",
        "Bond No 9","Boadicea","Carner Barcelona","Clean","Commodity",
        "Costume National","Creed","Derek Lam","Diptique","Estee Lauder",
        "Franck Olivier","Giorgio Beverly Hills","Guerlain","Guess",
        "Histoires de Parfums","Illuminum","Jimmy Choo","Kenneth Cole",
        "Lalique","Lolita Lempicka","Lubin","Miu Miu","Moresque",
        "Nobile 1942","Oscar de la Renta","Oud Elite","Philipp Plein",
        "Police","Prada","Rasasi","Reminiscence","Salvatore Ferragamo",
        "Stella McCartney","Ted Lapidus","Ungaro","Vera Wang","Viktor Rolf",
        "Zadig Voltaire","Zegna","Ajwad","Club de Nuit","Milestone",
        "لطافة","العربية للعود","رصاسي","أجمل","الحرمين","أرماف",
        "أمواج","كريد","توم فورد","ديور","شانيل","غوتشي","برادا",
        "ميسوني","جوسي كوتور","موسكينو","دانهيل","بنتلي",
        "كينزو","لاكوست","فندي","ايلي صعب","ازارو",
        "كيليان","نيشان","زيرجوف","بنهاليغونز","مارلي","جيرلان",
        "تيزيانا ترينزي","مايزون فرانسيس","بايريدو","لي لابو",
        "مانسيرا","مونتالي","روجا","جو مالون","ثمين","أمادو",
        "ناسوماتو","ميزون مارجيلا","نيكولاي",
        "جيمي تشو","لاليك","بوليس","فيكتور رولف",
        "كلوي","بالنسياغا","ميو ميو",
    ]
    WORD_REPLACEMENTS  = {}
    MATCH_THRESHOLD    = 85
    HIGH_CONFIDENCE    = 95
    PRICE_TOLERANCE    = 5
    TESTER_KEYWORDS    = ["tester", "تستر"]
    SET_KEYWORDS       = ["set", "طقم", "مجموعة"]
    GEMINI_API_KEYS    = []
    OPENROUTER_API_KEY = ""

# ─── مفاتيح Gemini: config أولاً (يدمج secrets.toml + env)؛ إن فارغ استخدم env فقط ───
import os as _os


def _load_gemini_keys_from_env():
    keys = []
    v = _os.environ.get("GEMINI_API_KEYS", "")
    if v:
        v = v.strip()
        if v.startswith("["):
            try:
                parsed = json.loads(v)
                if isinstance(parsed, list):
                    keys = [str(k).strip() for k in parsed if k]
            except Exception:
                keys += [k.strip() for k in v.split(",") if k.strip()]
        else:
            keys += [k.strip() for k in v.split(",") if k.strip()]
    # يدعم صيغتي الترقيم: GEMINI_KEY_N و GEMINI_API_KEY_N (حتى 30 مفتاحاً لتدويرها)
    for i in range(1, 31):
        for _pat in (f"GEMINI_API_KEY_{i}", f"GEMINI_KEY_{i}"):
            k = _os.environ.get(_pat, "")
            if k.strip():
                keys.append(k.strip())
    for env_name in ["GEMINI_API_KEY", "GEMINI_KEY"]:
        k = _os.environ.get(env_name, "")
        if k.strip():
            keys.append(k.strip())
    out = list(dict.fromkeys(keys))
    return [k for k in out if k and len(k) > 20]


try:
    GEMINI_API_KEYS
except NameError:
    GEMINI_API_KEYS = []

if not GEMINI_API_KEYS:
    GEMINI_API_KEYS = _load_gemini_keys_from_env()

# ─── مرادفات ذكية للعطور ────────────────────
_SYN = {
    "eau de parfum":"edp","او دو بارفان":"edp","أو دو بارفان":"edp",
    "او دي بارفان":"edp","بارفان":"edp","parfum":"edp","perfume":"edp",
    "eau de toilette":"edt","او دو تواليت":"edt","أو دو تواليت":"edt",
    "تواليت":"edt","toilette":"edt","toilet":"edt",
    "eau de cologne":"edc","كولون":"edc","cologne":"edc",
    "extrait de parfum":"extrait","parfum extrait":"extrait",
    "ديور":"dior","شانيل":"chanel","شنل":"chanel","أرماني":"armani","ارماني":"armani",
    "جورجيو ارماني":"armani","فرساتشي":"versace","فيرساتشي":"versace",
    "غيرلان":"guerlain","توم فورد":"tom ford","تومفورد":"tom ford",
    "لطافة":"lattafa","لطافه":"lattafa",
    "أجمل":"ajmal","رصاصي":"rasasi","أمواج":"amouage","كريد":"creed",
    "ايف سان لوران":"ysl","سان لوران":"ysl","yves saint laurent":"ysl",
    "غوتشي":"gucci","قوتشي":"gucci","برادا":"prada","برادة":"prada",
    "بربري":"burberry","بيربري":"burberry","جيفنشي":"givenchy","جفنشي":"givenchy",
    "كارولينا هيريرا":"carolina herrera","باكو رابان":"paco rabanne",
    "نارسيسو رودريغيز":"narciso rodriguez","كالفن كلاين":"calvin klein",
    "هوجو بوس":"hugo boss","فالنتينو":"valentino","بلغاري":"bvlgari",
    "كارتييه":"cartier","لانكوم":"lancome","جو مالون":"jo malone",
    "سوفاج":"sauvage","بلو":"bleu","إيروس":"eros","ايروس":"eros",
    "وان ميليون":"1 million",
    "إنفيكتوس":"invictus","أفينتوس":"aventus","عود":"oud","مسك":"musk",
    "ميسوني":"missoni","جوسي كوتور":"juicy couture","موسكينو":"moschino",
    "دانهيل":"dunhill","بنتلي":"bentley","كينزو":"kenzo","لاكوست":"lacoste",
    "فندي":"fendi","ايلي صعب":"elie saab","ازارو":"azzaro",
    "فيراغامو":"ferragamo","شوبار":"chopard","بوشرون":"boucheron",
    "لانكم":"lancome","لانكوم":"lancome","جيفنشي":"givenchy","جيفانشي":"givenchy",
    "بربري":"burberry","بيربري":"burberry","بوربيري":"burberry",
    "فيرساتشي":"versace","فرزاتشي":"versace",
    "روبيرتو كفالي":"roberto cavalli","روبرتو كافالي":"roberto cavalli",
    "سلفاتوري":"ferragamo","سالفاتوري":"ferragamo",
    "ايف سان لوران":"ysl","ايف سانت لوران":"ysl",
    "هيرميس":"hermes","ارميس":"hermes","هرمز":"hermes",
    "كيليان":"kilian","كليان":"kilian",
    "نيشان":"nishane","نيشاني":"nishane",
    "زيرجوف":"xerjoff","زيرجوفف":"xerjoff",
    "بنهاليغونز":"penhaligons","بنهاليغون":"penhaligons",
    "مارلي":"parfums de marly","دي مارلي":"parfums de marly",
    "جيرلان":"guerlain","غيرلان":"guerlain","جرلان":"guerlain",
    "تيزيانا ترينزي":"tiziana terenzi","تيزيانا تيرينزي":"tiziana terenzi",
    "تيزيانا":"tiziana terenzi","تيرينزي":"tiziana terenzi",
    "ناسوماتو":"nasomatto",
    "ميزون مارجيلا":"maison margiela","مارجيلا":"maison margiela","ربليكا":"replica",
    "نيكولاي":"nicolai","نيكولائي":"nicolai",
    "مايزون فرانسيس":"maison francis kurkdjian","فرانسيس":"maison francis kurkdjian",
    "بايريدو":"byredo","لي لابو":"le labo",
    "مانسيرا":"mancera","مونتالي":"montale","روجا":"roja",
    "جو مالون":"jo malone","جومالون":"jo malone",
    "ثمين":"thameen","أمادو":"amadou","امادو":"amadou",
    "انيشيو":"initio","إنيشيو":"initio","initio":"initio",
    "جيمي تشو":"jimmy choo","جيميتشو":"jimmy choo",
    "لاليك":"lalique","بوليس":"police",
    "فيكتور رولف":"viktor rolf","فيكتور اند رولف":"viktor rolf",
    "كلوي":"chloe","شلوي":"chloe",
    "بالنسياغا":"balenciaga","بالنسياجا":"balenciaga",
    "ميو ميو":"miu miu",
    "استي لودر":"estee lauder","استيلودر":"estee lauder",
    "كوتش":"coach","مايكل كورس":"michael kors",
    "رالف لورين":"ralph lauren","رالف لوران":"ralph lauren",
    "ايزي مياكي":"issey miyake","ايسي مياكي":"issey miyake",
    "دافيدوف":"davidoff","ديفيدوف":"davidoff",
    "دولشي اند غابانا":"dolce gabbana","دولتشي":"dolce gabbana","دولشي":"dolce gabbana",
    "جان بول غولتييه":"jean paul gaultier","غولتييه":"jean paul gaultier","غولتيه":"jean paul gaultier",
    "غوتييه":"jean paul gaultier","جان بول غوتييه":"jean paul gaultier","قوتييه":"jean paul gaultier","قولتييه":"jean paul gaultier",
    "مونت بلانك":"montblanc","مونتبلان":"montblanc",
    "موجلر":"mugler","موغلر":"mugler","تييري موجلر":"mugler",
    "كلوب دي نوي":"club de nuit","كلوب دنوي":"club de nuit",
    "مايلستون":"milestone",
    "سكاندل":"scandal","سكاندال":"scandal",
    " مل":" ml","ملي ":"ml ","ملي":"ml",
    "ليتر":"l","لتر":"l"," لتر":" l"," ليتر":" l",
    "جم":"g","جرام":"g"," غرام":" g",
    # ── توحيد الحروف العربية ──
    "أ":"ا","إ":"ا","آ":"ا","ة":"ه","ى":"ي","ؤ":"و","ئ":"ي","ـ":"",
    # ── تهجئات بديلة لكلمات العطور (الأهم للمطابقة) ──
    "بيرفيوم":"edp","بيرفيومز":"edp","بارفيومز":"edp","برفان":"edp",
    "پارفيوم":"edp","پرفيوم":"edp","بارفيم":"edp",
    "تواليت":"edt","تواليتة":"edt","طواليت":"edt",
    "اكسترايت":"extrait","اكستريت":"extrait","اكسترييت":"extrait",
    "انتينس":"intense","انتانس":"intense","إنتنس":"intense",
    # ── تهجئات الماركات الإضافية ──
    "ايسينشيال":"essential","اسنشيال":"essential","ايسانشيال":"essential",
    "اسنشال":"essential","ايسنشال":"essential","ايسينشال":"essential",
    "سولييل":"soleil","سولايل":"soleil","سوليل":"soleil",
    "فلورال":"floral","فلورل":"floral","فلوريل":"floral",
    "سوفاج":"sauvage","سوفايج":"sauvage","سافاج":"sauvage",
    "بلو":"bleu","بلوو":"bleu",
    "ليبر":"libre","ليبرة":"libre",
    "اوريجينال":"original","أوريجينال":"original",
    "إكسترا":"extra","اكسترا":"extra",
    "انفيوجن":"infusion","انفيجن":"infusion","انفيوزن":"infusion",
    "ديليت":"delight","ديلايت":"delight",
    "نيوتر":"neutre","نيوتره":"neutre","نيوتير":"neutre",
    "بيور":"pure","بيوره":"pure","بيورة":"pure",
    "نوار":"noir","نوير":"noir",
    "روز":"rose","روس":"rose",
    "جاسمين":"jasmine","جازمين":"jasmine","ياسمين":"jasmine",
    "ميلانجي":"melange","ميلانج":"melange",
    "بريلوج":"prelude","برولوج":"prelude",
    "ريزيرف":"reserve","ريزيرفي":"reserve",
    "اميثست":"amethyst","اميثيست":"amethyst",
    "دراكار":"drakkar","دراكر":"drakkar",
    "نمروود":"nimrod","نمرود":"nimrod",
    "اوليفيا":"olivia","اوليفيه":"olivia",
    "ليجند":"legend","ليجاند":"legend",
    "سبورت":"sport","سبورتس":"sport",
    "بلاك":"black","بلك":"black",
    "وايت":"white","وايث":"white",
    "جولد":"gold","قولد":"gold",
    "سيلفر":"silver","سيلفير":"silver",
    "نايت":"night","نايث":"night",
    # حُذفت: "داي"→day و"دي"→day لأنها حروف جر تُفسد أسماء الماركات (دي مارلي، دي جيو)
    # ── v26.0: مرادفات إضافية لزيادة الدقة ──
    # أحجام بديلة
    "٥٠":"50","٧٥":"75","١٠٠":"100","١٢٥":"125","١٥٠":"150","٢٠٠":"200",
    "٢٥٠":"250","٣٠٠":"300","٣٠":"30","٨٠":"80",
    # تركيزات إضافية
    "بارفيوم انتنس":"edp intense","انتنس":"intense","إنتنس":"intense",
    "ابسولو":"absolue","ابسوليو":"absolue","ابسوليوت":"absolute",
    "اكستريم":"extreme","اكسترييم":"extreme",
    "بريفيه":"prive","بريفي":"prive","privee":"prive","privé":"prive",
    "ليجير":"legere","ليجيره":"legere","légère":"legere",
    # ماركات ناقصة
    "توماس كوسمالا":"thomas kosmala","كوسمالا":"thomas kosmala",
    "روسيندو ماتيو":"rosendo mateu","ماتيو":"rosendo mateu",
    "بوديسيا":"boadicea","بواديسيا":"boadicea",
    "نوبيلي":"nobile","نوبيل":"nobile",
    "كارنر":"carner","كارنير":"carner",
    "اتيليه كولون":"atelier cologne","اتيليه":"atelier",
    "بوند نمبر ناين":"bond no 9","بوند":"bond",
    "هيستوار":"histoires","هيستوريز":"histoires",
    "لوبين":"lubin","لوبان":"lubin",
    "فيليب بلين":"philipp plein","فيلب بلين":"philipp plein",
    "اوسكار دي لا رنتا":"oscar de la renta","اوسكار":"oscar",
    "ستيلا مكارتني":"stella mccartney","ستيلا":"stella",
    "زاديغ":"zadig","زاديج":"zadig",
    "تيد لابيدوس":"ted lapidus","لابيدوس":"ted lapidus",
    "انقارو":"ungaro","اونغارو":"ungaro",
    "فيرا وانق":"vera wang","فيرا وانغ":"vera wang",
    "كينيث كول":"kenneth cole","كينث كول":"kenneth cole",
    "اد هاردي":"ed hardy","ايد هاردي":"ed hardy",
    "دنهل":"dunhill","دنهيل":"dunhill","دانهيل":"dunhill","الفريد دنهل":"alfred dunhill",
    "بنتلي":"bentley","بنتلى":"bentley",
    "جاغوار":"jaguar","جاكوار":"jaguar",
    # كلمات عطرية شائعة
    "عنبر":"amber","عنبري":"amber","امبر":"amber",
    "عود":"oud","عودي":"oud",
    "مسك":"musk","مسكي":"musk","موسك":"musk",
    "زعفران":"saffron","زعفراني":"saffron",
    "بخور":"incense","بخوري":"incense",
    "فانيلا":"vanilla","فانيليا":"vanilla",
    "باتشولي":"patchouli",
    "صندل":"sandalwood","صندلي":"sandalwood",
    "توباكو":"tobacco","تبغ":"tobacco",
    # تصحيح إملائي شائع
    "بيرفوم":"edp","بريفيوم":"edp","بارفوم":"edp",
    "تولت":"edt","تويلت":"edt",
}

# ─── ⚡ v31: regex مُجمَّع للمرادفات — يُحوِّل 280 str.replace إلى عملية واحدة ───
_SYN_SORTED = sorted(_SYN.keys(), key=len, reverse=True)  # الأطول أولاً لمنع تعارض
_SYN_RE = re.compile("|".join(re.escape(k) for k in _SYN_SORTED), re.UNICODE)
def _syn_replace(t):
    """استبدال المرادفات بـ regex واحد بدل 280 حلقة — ~10x أسرع"""
    return _SYN_RE.sub(lambda m: _SYN[m.group(0)], t)

# ─── v22: Dynamic Brand Vocabulary Enrichment ────────────────
_brands_enriched = False

def enrich_known_brands(comp_dfs=None, db_path=None):
    """
    إثراء KNOWN_BRANDS بماركات المنافسين الحقيقية من competitor_products_store.
    يُستدعى مرة واحدة عند بدء التحليل — يرفع التغطية من 46% إلى 92%+.
    """
    global KNOWN_BRANDS, _brands_enriched
    if _brands_enriched:
        return

    new_brands = set()

    # Source 1: brand column from competitor DataFrames
    if comp_dfs:
        for cdf in comp_dfs.values():
            for col in ("brand", "الماركة"):
                if col in cdf.columns:
                    vals = cdf[col].fillna("").astype(str).str.strip()
                    new_brands.update(v for v in vals if v and len(v) >= 2 and v.lower() not in ("nan", "none", "0"))
                    break

    # Source 2: DB if available
    if db_path:
        try:
            import sqlite3
            conn = sqlite3.connect(db_path)
            rows = conn.execute("SELECT DISTINCT brand FROM competitor_products_store WHERE brand IS NOT NULL AND brand != ''").fetchall()
            conn.close()
            new_brands.update(r[0].strip() for r in rows if r[0] and len(r[0].strip()) >= 2)
        except Exception:
            pass

    if not new_brands:
        _brands_enriched = True
        return

    # Normalize and deduplicate against existing KNOWN_BRANDS
    existing_lower = {b.lower() for b in KNOWN_BRANDS}
    added = 0
    for nb in new_brands:
        # Clean: remove very long strings, numbers-only, URLs
        if len(nb) > 50 or nb.startswith("http") or nb.isdigit():
            continue
        if nb.lower() not in existing_lower:
            KNOWN_BRANDS.append(nb)
            existing_lower.add(nb.lower())
            added += 1

    # Clear LRU caches so they rebuild with enriched list
    _get_normalized_brands.cache_clear()
    _get_brand_normalized_pairs.cache_clear()

    _brands_enriched = True
    import logging
    logging.getLogger("engines.engine").info(
        "Brand vocabulary enriched: +%d brands (total: %d)", added, len(KNOWN_BRANDS))


# ─── v26.0: Fuzzy Spell Correction ────────────────
# ✅ إصلاح #5: LRU cache لتجنب إعادة بناء قائمة الماركات مع كل استدعاء
@_functools.lru_cache(maxsize=1)
def _get_normalized_brands():
    """يُحسب مرة واحدة فقط عند أول استدعاء"""
    return [(b, b.lower()) for b in KNOWN_BRANDS]


@_functools.lru_cache(maxsize=1)
def _get_brand_normalized_pairs():
    """⚡ v31: أزواج (ماركة أصلية, ماركة مطبَّعة, ماركة lower) — تُحسب مرة واحدة"""
    pairs = []
    for b in KNOWN_BRANDS:
        n = b.strip().lower()
        # تطبيع خفيف بدون normalize() الكاملة لتجنب تغيير اسم الماركة
        n = _syn_replace(n)
        pairs.append((b, n, b.lower()))
    return pairs


def _fuzzy_correct_brand(text: str, threshold: int = 82) -> str:
    """تصحيح إملائي ذكي للماركات — يُستخدم عند فشل المطابقة المباشرة"""
    if not text or len(text) < 3:
        return ""
    from rapidfuzz import fuzz as _fz
    text_norm = text.lower().strip()
    best_brand = ""
    best_score = 0
    for b_orig, b_low in _get_normalized_brands():
        s = _fz.ratio(text_norm, b_low)
        if s > best_score and s >= threshold:
            best_score = s
            best_brand = b_orig
            if best_score >= 97:
                break
    return best_brand

# ─── SQLite Cache — اتصال دائم مع thread safety ────────────────
# ✅ إصلاح #6: اتصال دائم بدلاً من فتح/إغلاق لكل عملية
_DB = get_data_db_path("match_cache_v22.db")
_db_conn = None
_db_lock = threading.Lock()


def _get_db_conn():
    """يُعيد اتصالاً دائماً — يُنشئه عند أول استدعاء فقط"""
    global _db_conn
    with _db_lock:
        if _db_conn is None:
            try:
                _db_conn = sqlite3.connect(_DB, check_same_thread=False)
                _db_conn.execute(
                    "CREATE TABLE IF NOT EXISTS cache(h TEXT PRIMARY KEY, v TEXT, ts TEXT)"
                )
                _db_conn.execute("PRAGMA journal_mode=WAL")
                _db_conn.commit()
            except Exception:
                _db_conn = None
        return _db_conn


def _init_db():
    try:
        _get_db_conn()
    except Exception:
        pass


def _cget(k):
    try:
        conn = _get_db_conn()
        if conn is None:
            return None
        with _db_lock:
            r = conn.execute("SELECT v FROM cache WHERE h=?", (k,)).fetchone()
        return json.loads(r[0]) if r else None
    except Exception:
        return None


def _cset(k, v):
    try:
        conn = _get_db_conn()
        if conn is None:
            return
        with _db_lock:
            conn.execute(
                "INSERT OR REPLACE INTO cache VALUES(?,?,?)",
                (k, json.dumps(v, ensure_ascii=False), datetime.now().isoformat())
            )
            conn.commit()
    except Exception:
        pass


_init_db()

# ─── استبعاد أعمدة كشط (أسماء CSS / Tailwind / حروف عشوائية) ─────────────
_SCRAPER_COL_RE = re.compile(
    r"(^|\s)(w-|h-|p-|m-|text-|bg-|flex|grid|gap-|rounded|sm:|md:|lg:|xl:|"
    r"hover:|focus:|items-|justify-|min-w|max-w|truncate|styles_|productCard|"
    r"__|src\b|cls\b|className|w-full|h-full)(\s|$)",
    re.I | re.UNICODE,
)


def _is_scraper_column_name(col):
    s = str(col).strip()
    if not s:
        return True
    if _SCRAPER_COL_RE.search(s):
        return True
    if s.count("__") >= 2 and len(s) > 24:
        return True
    if "style" in s.lower() and "__" in s:
        return True
    return False


def _drop_scraper_columns(df):
    """حذف أعمدة تبدو كمخرجات كشط وليست حقولاً حقيقية."""
    if df is None or df.empty:
        return df
    keep = [c for c in df.columns if not _is_scraper_column_name(c)]
    if not keep:
        return df
    if len(keep) < len(df.columns):
        return df[keep].copy()
    return df


def _normalize_header_typos(df):
    """توحيد أشهر أخطاء التصدير (أسم → اسم، صوره → صورة)."""
    if df is None or df.empty:
        return df
    m = {}
    for c in df.columns:
        ns = str(c).strip().replace("\ufeff", "")
        ns = ns.replace("أسم المنتج", "اسم المنتج").replace("أسم ", "اسم ")
        ns = ns.replace("صوره المنتج", "صورة المنتج").replace("صوره ", "صورة ")
        if ns != str(c).strip():
            m[c] = ns
    if m:
        df = df.rename(columns=m)
    return df


_IMG_URL_RE = re.compile(r"\.(webp|jpg|jpeg|png|gif|avif|svg)(\?|#|\"|'|$)", re.I)


def _looks_like_image_url(s: str) -> bool:
    """رابط يبدو ملف صورة (للاحتياط عندما لا يُعرف عمود الصورة بالاسم)."""
    if not s:
        return False
    vl = s.strip().lower()
    if "http" not in vl and not vl.startswith("//"):
        # مسار نسبي شائع في التصديرات
        if vl.startswith("/") and any(x in vl for x in (".jpg", ".jpeg", ".png", ".webp", ".gif", ".avif")):
            return True
        return False
    if _IMG_URL_RE.search(vl):
        return True
    return any(x in vl for x in (".jpg", ".jpeg", ".png", ".webp", ".gif", ".avif"))


_EMBEDDED_HTTP_IMG = re.compile(
    r'https?://[^\s<>"\'\)]+\.(?:webp|jpg|jpeg|png|gif|avif)(?:\?[^\s<>"\'\)\]]*)?',
    re.I,
)


def _extract_image_url_from_cell(val) -> str:
    """خلية مباشرة أو نص/HTML (وصف صورة، src=...) يضم رابط صورة."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    s = first_image_url_string(str(val).strip())
    if not s or s.lower() in ("nan", "none", "<na>"):
        return ""
    if _looks_like_image_url(s):
        return s.split()[0]
    m = _EMBEDDED_HTTP_IMG.search(s)
    if m:
        return m.group(0).strip().rstrip(".,;)]\"'")
    m2 = re.search(
        r'https?://[^\s<>"\'\)]+/(?:images?|img|media|storage|uploads|files|cdn)/[^\s<>"\'\)]+',
        s,
        re.I,
    )
    if m2:
        t = m2.group(0).strip().rstrip(".,;)]\"'")
        if len(t) < 800:
            return t
    return ""


def _column_content_scores(series):
    """نِسَب: روابط http، صور، أسعار."""
    vals = series.dropna().head(60).astype(str)
    n = len(vals)
    if n == 0:
        return 0.0, 0.0, 0.0
    http_n = img_n = price_n = 0
    for v in vals:
        vl = v.strip().lower()
        if "http://" in vl or "https://" in vl or vl.startswith("//"):
            http_n += 1
        if _IMG_URL_RE.search(vl) or ("http" in vl and any(
            x in vl for x in (".jpg", ".png", ".webp", ".jpeg", ".gif"))):
            img_n += 1
        try:
            x = float(str(v).replace(",", "").replace("ر.س", "").replace("﷼", "").strip())
            if 0.5 <= x <= 800000:
                price_n += 1
        except (ValueError, TypeError):
            pass
    return http_n / n, img_n / n, price_n / n


def _infer_column_roles(df):
    """
    بعد تطبيع الأسماء: إن بقيت أعمدة غير معروفة، خمّنها من المحتوى
    (روابط، صور، أسعار، اسم المنتج).
    """
    if df is None or df.empty:
        return df
    cols = list(df.columns)

    def _has(col_name):
        return any(str(c).strip() == col_name for c in cols)

    has_name = _has("اسم المنتج") or any(str(c).strip() == "المنتج" for c in cols)
    has_price = any("سعر" in str(c) for c in cols) or _has("السعر")
    has_img = _has("صورة المنتج")
    has_link = _has("رابط المنتج")

    scored = []
    for c in cols:
        sc = str(c).strip()
        if sc in ("اسم المنتج", "المنتج", "سعر المنتج", "السعر", "صورة المنتج",
                  "رابط المنتج", "رمز المنتج sku", "رمز المنتج", "No.", "no."):
            continue
        http_r, img_r, price_r = _column_content_scores(df[c])
        scored.append((c, http_r, img_r, price_r))

    rename = {}
    # صورة: أعلى نسبة امتدادات صور / روابط صور
    if not has_img:
        scored_img = sorted(scored, key=lambda x: -x[2])
        for c, hr, ir, pr in scored_img:
            if ir >= 0.22 and ir >= hr * 0.35:
                rename[c] = "صورة المنتج"
                has_img = True
                break

    # رابط صفحة المنتج (http بدون سيطرة صورة)
    if not has_link:
        scored_http = sorted(scored, key=lambda x: -x[1])
        for c, hr, ir, pr in scored_http:
            if c in rename:
                continue
            if hr >= 0.38 and ir < 0.55:
                rename[c] = "رابط المنتج"
                has_link = True
                break

    # سعر
    if not has_price:
        scored_price = sorted(scored, key=lambda x: -x[3])
        for c, hr, ir, pr in scored_price:
            if c in rename:
                continue
            if pr >= 0.5:
                rename[c] = "سعر المنتج"
                break

    # اسم
    if not has_name:
        for c, hr, ir, pr in scored:
            if c in rename:
                continue
            if pr < 0.35 and hr < 0.25 and ir < 0.2:
                txt = " ".join(df[c].dropna().head(5).astype(str))
                if len(txt) >= 20:
                    rename[c] = "اسم المنتج"
                    break

    if rename:
        df = df.rename(columns=rename)
    return df


# ─── دوال أساسية ────────────────────────────
def read_file(f):
    try:
        name = f.name.lower()
        df = None
        if name.endswith('.csv'):
            for enc in ['utf-8-sig','utf-8','windows-1256','cp1256','latin-1']:
                try:
                    f.seek(0)
                    peek = pd.read_csv(f, header=None, nrows=6, encoding=enc, on_bad_lines='skip')
                    f.seek(0)
                    use_row1 = _should_use_second_row_header(peek)
                    f.seek(0)
                    if use_row1:
                        df = pd.read_csv(f, header=1, encoding=enc, on_bad_lines='skip')
                    else:
                        df = pd.read_csv(f, encoding=enc, on_bad_lines='skip')
                    if len(df) > 0 and not str(df.columns[0]).startswith('\ufeff'):
                        break
                except Exception:
                    continue
            if df is None:
                return None, "فشل قراءة الملف بجميع الترميزات"
        elif name.endswith(('.xlsx','.xls')):
            f.seek(0)
            peek = pd.read_excel(f, header=None, nrows=4, engine=None)
            f.seek(0)
            if _should_use_second_row_header(peek):
                df = pd.read_excel(f, header=1)
            else:
                df = pd.read_excel(f)
        else:
            return None, "صيغة غير مدعومة"
        df.columns = df.columns.map(lambda x: str(x).strip().replace('\ufeff', ''))
        df = df.dropna(how='all').reset_index(drop=True)
        df = _normalize_header_typos(df)
        # إعادة تسمية أعمدة الكشط/سلة *قبل* حذف الأعمدة المشبوهة — وإلا تُفقد الحقول الأساسية
        df = _detect_double_header(df)
        df = _smart_rename_columns(df)
        df = _drop_scraper_columns(df)
        df = _infer_column_roles(df)
        return df, None
    except Exception as e:
        return None, str(e)


def _should_use_second_row_header(peek):
    """ملف سلة/متجر: الصف 0 مجموعات (مثل بيانات المنتج)، الصف 1 عناوين الحقول."""
    if peek is None or len(peek) < 2:
        return False
    row0 = [str(x).strip().lower() for x in peek.iloc[0].tolist()]
    row1 = [str(x).strip() for x in peek.iloc[1].tolist()]
    row0_join = " ".join(row0)
    keys1 = (
        "اسم المنتج", "أسم المنتج", "سعر المنتج", "صورة المنتج", "رابط المنتج", "رابط",
        "تصنيف المنتج", "وصف صورة المنتج", "نوع المنتج", "الكمية المتوفرة",
        "no.", "no", "النوع", "رمز المنتج", "sku", "product", "name", "price",
        "سعر التكلفة", "السعر المخفض", "باركود", "الوصف", "الماركة",
    )
    hits1 = sum(1 for x in row1 if any(k in x.lower() for k in keys1))
    hits0 = sum(1 for x in row0 if any(k in x for k in keys1))
    non_empty0 = sum(1 for x in row0 if x and x != "nan")
    group_like = sum(1 for x in row0 if "بيانات" in x or ("منتج" in x and len(x) < 40) or x == "")
    # صف علوي يعبّر عن مجموعة حقول (تصدير سلة / Excel)
    if "بيانات المنتج" in row0_join or "بيانات" in row0_join:
        if hits1 >= 2:
            return True
    if hits1 >= 3 and hits1 >= hits0:
        return True
    if hits1 >= 2 and group_like >= max(1, non_empty0 // 2):
        return True
    return False


def _detect_double_header(df):
    """كشف ملفات ذات صفين عناوين (مثل ملف سلة الذي يحتوي على صف مجموعة + صف عناوين)"""
    cols = list(df.columns)
    unnamed_count = sum(1 for c in cols if str(c).startswith('Unnamed'))
    # إذا أغلب الأعمدة Unnamed → الصف الأول من البيانات قد يكون العناوين الحقيقية
    if unnamed_count >= len(cols) // 2 and len(df) > 2:
        # تحقق: هل الصف الأول يحتوي على أسماء أعمدة معروفة؟
        first_row = df.iloc[0].astype(str).tolist()
        _known_headers = [
            'اسم المنتج', 'أسم المنتج', 'سعر المنتج', 'السعر', 'النوع',
            'صورة المنتج', 'رابط المنتج', 'وصف صورة المنتج', 'تصنيف المنتج', 'نوع المنتج',
            'no.', 'no', 'sku', 'رمز المنتج', 'سعر التكلفة', 'السعر المخفض',
            'product', 'name', 'price', 'رقم المنتج', 'رمز المنتج sku', 'الكمية المتوفرة',
            'الباركود', 'الماركة',
        ]

        def _hdr_hit(cell):
            s = str(cell).strip().lower().replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
            for h in _known_headers:
                hn = h.lower().replace("أ", "ا")
                if s == hn or hn in s or s in hn:
                    return True
            return False

        match_count = sum(1 for v in first_row if _hdr_hit(v))
        if match_count >= 2:
            # الصف الأول هو العناوين الحقيقية → استخدمه كعناوين
            new_cols = [str(v).strip() for v in first_row]
            df.columns = new_cols
            df = df.iloc[1:].reset_index(drop=True)
    return df


def _smart_rename_columns(df):
    """التعرف العميق على الأعمدة (Scraper CSS + محتوى) — أسماء موحّدة مع _fcol و resolve_catalog_columns."""
    if df is None or df.empty:
        return df
    cols = list(df.columns)

    is_dirty = any(
        "__" in str(c)
        or "style" in str(c).lower()
        or "productcard" in str(c).lower()
        or "text-" in str(c).lower()
        or "w-full" in str(c).lower()
        or "abs-" in str(c).lower()
        or "href" in str(c).lower()
        or "src" in str(c).lower()
        or str(c).lower().startswith("unnamed")
        for c in cols
    )

    def _clean_arabic_headers():
        blob = " ".join(str(c) for c in cols).lower()
        return ("اسم" in blob or "منتج" in blob) and ("سعر" in blob or "price" in blob)

    if len(cols) == 4 and not is_dirty and _clean_arabic_headers():
        return df

    if not is_dirty and len(cols) != 4:
        return df

    # أنماط شائعة في تصديرات الكشط (الأكثر تحديداً أولاً)
    # — تغطي: CSS class names / HTML attrs / English / Arabic keywords
    CSS_PATTERNS = [
        # ── CSS class names (worldgivenchy, saeedsalah, …) ──
        ("styles_productcard__name",  "اسم المنتج"),
        ("productcard__name",         "اسم المنتج"),
        ("text-sm-2",                 "سعر المنتج"),
        ("text-sm",                   "سعر المنتج"),
        ("abs-size href",             "رابط المنتج"),
        ("w-full src",                "صورة المنتج"),
        ("w-full",                    "صورة المنتج"),
        # ── HTML attributes (standalone) ──
        ("href",                      "رابط المنتج"),
        ("src",                       "صورة المنتج"),
        # ── English keywords ──
        ("product_name",              "اسم المنتج"),
        ("productname",               "اسم المنتج"),
        ("product_title",             "اسم المنتج"),
        ("title",                     "اسم المنتج"),
        ("price",                     "سعر المنتج"),
        ("image_url",                 "صورة المنتج"),
        ("image",                     "صورة المنتج"),
        ("img",                       "صورة المنتج"),
        ("photo",                     "صورة المنتج"),
        ("product_url",               "رابط المنتج"),
        ("product_link",              "رابط المنتج"),
        ("link",                      "رابط المنتج"),
        ("url",                       "رابط المنتج"),
        ("name",                      "اسم المنتج"),
        # ── Arabic keywords ──
        ("اسم",                       "اسم المنتج"),
        ("سعر",                       "سعر المنتج"),
        ("صورة",                      "صورة المنتج"),
        ("صوره",                      "صورة المنتج"),
        ("رابط",                      "رابط المنتج"),
    ]

    KNOWN_EXACT = frozenset({
        "اسم المنتج", "أسم المنتج", "المنتج", "سعر المنتج", "السعر", "سعر",
        "صورة المنتج", "صوره المنتج", "رابط المنتج", "الرابط", "الماركة",
        "رقم المنتج", "معرف المنتج", "رمز المنتج", "رمز المنتج sku",
    })
    KNOWN_EXACT_EN = frozenset({
        "product name", "product_name", "name", "price", "sku", "title",
        "product id", "product_id", "link", "url", "image",
    })

    def _known_header(c):
        s = str(c).strip()
        if s in KNOWN_EXACT:
            return True
        sl = s.lower().replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
        if sl in KNOWN_EXACT_EN:
            return True
        return False

    new_cols = {}
    used = set()

    for col in cols:
        if _known_header(col):
            continue
        csl = str(col).lower()
        for needle, std in CSS_PATTERNS:
            if needle in csl:
                if std in used:
                    continue
                new_cols[col] = std
                used.add(std)
                break

    for col in cols:
        if col in new_cols or _known_header(col):
            continue
        c_str = str(col).strip()
        need_heuristic = (
            c_str.startswith("Unnamed")
            or "__" in c_str
            or "style" in c_str.lower()
            or "text-" in c_str.lower()
            or "href" in c_str.lower()
            or "src" in c_str.lower()
            or "w-full" in c_str.lower()
        )
        if not need_heuristic:
            continue

        sample = df[col].dropna().astype(str).head(30)
        if sample.empty:
            continue
        vs = [v.strip() for v in sample.tolist()]
        n = len(vs)

        numeric_count = 0
        for v in vs:
            try:
                x = float(
                    v.replace(",", "")
                    .replace("ر.س", "")
                    .replace("﷼", "")
                    .replace("SAR", "")
                    .strip()
                )
                if 0 <= x <= 10_000_000:
                    numeric_count += 1
            except (ValueError, TypeError):
                pass
        if numeric_count >= n * 0.6 and "سعر المنتج" not in used:
            new_cols[col] = "سعر المنتج"
            used.add("سعر المنتج")
            continue

        url_count = sum(1 for v in vs if v.startswith("http"))
        if url_count >= n * 0.5:
            img_count = sum(
                1
                for v in vs
                if (
                    ("cdn.salla" in v or "cdn." in v.lower())
                    or ".jpg" in v.lower()
                    or ".png" in v.lower()
                    or ".webp" in v.lower()
                    or ".jpeg" in v.lower()
                    or _IMG_URL_RE.search(v.lower())
                )
            )
            if img_count >= max(1, n * 0.4) and "صورة المنتج" not in used:
                new_cols[col] = "صورة المنتج"
                used.add("صورة المنتج")
            elif "رابط المنتج" not in used:
                new_cols[col] = "رابط المنتج"
                used.add("رابط المنتج")
            continue

        if "اسم المنتج" not in used:
            new_cols[col] = "اسم المنتج"
            used.add("اسم المنتج")
        else:
            new_cols[col] = col

    if new_cols:
        df = df.rename(columns=new_cols)

    # تنظيف إلزامي للأعمدة الأساسية الأربعة — يمنع NaN من كسر المحرك لاحقاً
    for _req_col, _is_url in [
        ("اسم المنتج",  False),
        ("سعر المنتج",  False),
        ("صورة المنتج", True),
        ("رابط المنتج", True),
    ]:
        if _req_col in df.columns:
            df[_req_col] = (
                df[_req_col]
                .fillna("")
                .astype(str)
                .str.strip()
                # تنظيف حرف الاقتباس والفراغات الملتصقة بالروابط
                .str.strip('"\'') if _is_url
                else df[_req_col].fillna("").astype(str).str.strip()
            )

    return df

# ── كلمات الضجيج التي تُشوّش المطابقة ──────────────────────────────
_NOISE_RE = re.compile(
    r'\b(عطر|تستر|تيستر|tester|'
    r'بارفيوم|بيرفيوم|بارفيومز|بيرفيومز|برفيوم|برفان|بارفان|بارفيم|'
    r'تواليت|تواليتة|كولون|اكسترايت|اكستريت|اكسترييت|'
    r'او\s*دو|او\s*دي|أو\s*دو|أو\s*دي|'
    r'الرجالي|النسائي|للجنسين|رجالي|نسائي|'
    r'parfum|perfume|cologne|toilette|extrait|intense|'
    r'eau\s*de|pour\s*homme|pour\s*femme|for\s*men|for\s*women|unisex|'
    r'edp|edt|edc)\b'
    r'|\b\d+(?:\.\d+)?\s*(?:ml|مل|ملي|oz)\b'   # أحجام: 100ml, 50مل
    r'|\b(100|200|50|75|150|125|250|300|30|80)\b',  # أرقام أحجام منفردة
    re.UNICODE | re.IGNORECASE
)

def normalize(text):
    """تطبيع قياسي: يوحّد الحروف والمرادفات مع الحفاظ على كامل النص"""
    if not isinstance(text, str): return ""
    t = text.strip().lower()
    # ⚡ v31: regex مُجمَّع بدل 280 str.replace — ~10x أسرع
    t = _syn_replace(t)
    # 2. توحيد الهمزات بعد المرادفات (لمعالجة ما تبقى من نصوص)
    for src, dst in [('أ','ا'),('إ','ا'),('آ','ا'),('ة','ه'),
                     ('ى','ي'),('ؤ','و'),('ئ','ي'),('ـ','')]:
        t = t.replace(src, dst)
    # 3. المرادفات المخصصة (من config.py)
    for k, v in WORD_REPLACEMENTS.items():
        t = t.replace(k.lower(), v)
    t = re.sub(r'[^\w\s\u0600-\u06FF.]', ' ', t)
    return re.sub(r'\s+', ' ', t).strip()


def normalize_name(text):
    """
    الدالة الموحدة للمطابقة — تُستخدم حصراً لمقارنة الأسماء.
    تحذف: عطر/بارفيوم/بيرفيوم/تستر/مل/edp/edt/للجنسين/100/50/...
    توحّد: أ/إ/آ→ا  ة/ه→ه  ى→ي
    المثال: 'عطر ايسينشيال بيرفيوم فيج انفيوجن 100مل' → 'essential فيج infusion'
    """
    if not isinstance(text, str): return ""
    t = text.strip().lower()
    # ⚡ v31: regex مُجمَّع بدل 280 str.replace
    t = _syn_replace(t)
    # 2. توحيد الهمزات بعد المرادفات
    for src, dst in [('أ','ا'),('إ','ا'),('آ','ا'),('ة','ه'),
                     ('ى','ي'),('ؤ','و'),('ئ','ي'),('ـ','')]:
        t = t.replace(src, dst)
    # 3. حذف كلمات الضجيج
    t = _NOISE_RE.sub(' ', t)
    # 4. v31.6: حذف الأرقام المتبوعة بوحدة قياس فقط (حماية أرقام المنتج مثل 212, 360)
    t = re.sub(r'\b\d+\s*(?:ml|مل|g|جم|جرام|oz|اونس|l|لتر)\b', ' ', t, flags=re.IGNORECASE)
    t = re.sub(r'[^\w\s\u0600-\u06FF]', ' ', t)
    return re.sub(r'\s+', ' ', t).strip()


# alias للتوافق مع الكود القديم
normalize_aggressive = normalize_name

def extract_size(text):
    if not isinstance(text, str): return 0.0
    tl = text.lower()
    # البحث عن oz أولاً وتحويله لـ ml
    oz = re.findall(r'(\d+(?:\.\d+)?)\s*(?:fl\.?\s*oz|oz|ounce|fluid\s*oz)', tl)
    if oz:
        return float(oz[0]) * 29.5735  # 1 oz = 29.5735 ml
    # البحث عن ml
    ml = re.findall(r'(\d+(?:\.\d+)?)\s*(?:ml|مل|ملي|milliliter)', tl)
    return float(ml[0]) if ml else 0.0

@_functools.lru_cache(maxsize=20000)
def extract_brand_fast(text):
    """ماركة سريعة: المرحلة المباشرة فقط (بدون تصحيح إملائي ضبابي).
    ~0.5ms/استدعاء مقابل ~6ms لـ extract_brand — للحجب الجماعي (المفقودات).
    تُعيد ماركة KNOWN_BRANDS قانونية (متّسقة على الطرفين) أو "".
    """
    if not isinstance(text, str) or not text:
        return ""
    n = normalize(text)
    tl = text.lower()
    for b_orig, b_norm, b_low in _get_brand_normalized_pairs():
        if b_norm in n or b_low in tl:
            return b_orig
    return ""


@_functools.lru_cache(maxsize=2000)
def extract_brand(text):
    if not isinstance(text, str): return ""
    n = normalize(text)
    tl = text.lower()
    # 1. ⚡ v31: مطابقة مباشرة مع أزواج مطبَّعة مسبقاً (بدل normalize لكل ماركة)
    for b_orig, b_norm, b_low in _get_brand_normalized_pairs():
        if b_norm in n or b_low in tl: return b_orig
    # 2. v26.0: تصحيح إملائي ذكي (fallback)
    words = text.split()
    # ✅ إصلاح #5: حد أقصى 12 مجموعة كلمات لتجنب O(N×W×B) التكرار المفرط
    candidates_checked = 0
    for i in range(len(words)):
        for length in [3, 2, 1]:
            if i + length <= len(words):
                candidate = " ".join(words[i:i+length])
                if len(candidate) >= 4:
                    corrected = _fuzzy_correct_brand(candidate, threshold=85)
                    if corrected:
                        return corrected
                    candidates_checked += 1
                    if candidates_checked >= 12:
                        return ""
    return ""

def extract_type(text):
    """
    v34: استخراج تركيز العطر بدقة من النص الخام (عربي+إنجليزي) — قبل أن يشوّهه normalize.
    يُرجع رمزاً موحّداً: PARFUM / EDP / EDT / EDC / ELIXIR / EXCLUSIF / FRAICHE  (+ '+INT' للنسخ المكثّفة)
    قاعدة SKU: اختلاف التركيز = منتج مختلف. (شانيل بلو اكسكلوسيف ≠ EDP ≠ EDT)
    """
    if not isinstance(text, str): return ""
    t = text.lower()
    # توحيد بسيط للهمزات/الياء دون تحويل لاتيني (نعمل على الخام)
    t = (t.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
           .replace("ى", "ي").replace("ـ", ""))
    base = ""
    # ── الأكثر تخصصاً أولاً ──
    if any(k in t for k in ("exclusif", "exclusive", "اكسكلوسيف", "اكسكلوزيف",
                            "ليكسكلوسيف", "لكسكلوسيف", "الحصري", " حصري")):
        base = "EXCLUSIF"
    elif any(k in t for k in ("elixir", "اليكسير", "الاكسير", "اكسير", "إكسير")):
        base = "ELIXIR"
    elif any(k in t for k in ("extrait", "اكستريت", "اكسترايت", "اكستريه", "extract de parfum",
                              "بيور بارفان", "بيور بارفيوم", "pure parfum", "le parfum",
                              "خلاصة العطر", "اكستريت دو بارفان", "اكسترايت دو",
                              "كونسنتري", "كونسنتريه", "كونسنتريت", "concentree", "concentrated",
                              "concentre", "مركز", "المركز")):
        base = "PARFUM"
    elif any(k in t for k in ("eau de parfum", "edp", "e.d.p", "او دو بارفيوم", "او دو برفيوم",
                              "او دي بارفيوم", "او دي برفيوم", "ادو بارفيوم", "أو دو بارفيوم",
                              "او دو بارفان", "او دو برفان", "اي دي بارفيوم", "اي دي برفيوم",
                              " بارفيوم", " برفيوم")):
        base = "EDP"
    elif any(k in t for k in ("eau de toilette", "edt", "e.d.t", "او دو تواليت", "او دي تواليت",
                              "ادو تواليت", "او دو توالت", "اي دي تواليت", "اي دي تويليت",
                              "او دو تويليت", "تويليت", "تويلت", " تواليت", " توالت")):
        base = "EDT"
    elif any(k in t for k in ("eau de cologne", "edc", "cologne", "كولونيا", "او دو كولونيا",
                              "او دو كولون", "كولون")):
        base = "EDC"
    elif any(k in t for k in ("eau fraiche", "fraiche", "fraîche", "او فريش", "فريش", "اقوا")):
        base = "FRAICHE"
    elif "parfum" in t and "shop" not in t:
        # "parfum" مجرّدة بالإنجليزية → غالباً eau de parfum
        base = "EDP"
    # ── معدِّل التكثيف (Intense) — نسخة مختلفة من نفس التركيز ──
    if base and any(k in t for k in ("intense", "انتنس", "انتينس", " مكثف")):
        base = base + "+INT"
    return base

def extract_gender(text):
    if not isinstance(text, str): return ""
    tl = text.lower()
    # تم التحديث ليشمل mans وصيغ الرجال المطلوبة
    m = any(k in tl for k in ["pour homme","for men","for him"," men "," men"," man ","رجالي","للرجال"," مان "," هوم ","homme"," uomo", "mans", "for mans", " mans "])
    w = any(k in tl for k in ["pour femme","for women","for her","women"," woman ","نسائي","للنساء","النسائي","lady","femme"," donna"])
    if m and not w: return "رجالي"
    if w and not m: return "نسائي"
    return ""

def extract_product_line(text, brand=""):
    """استخراج اسم خط الإنتاج (المنتج الأساسي) بعد إزالة الماركة والكلمات الشائعة.
    مثال: 'عطر بربري هيرو أو دو تواليت 100مل' → 'هيرو'
    مثال: 'عطر لندن من بربري للرجال' → 'لندن'
    هذا ضروري لمنع مطابقة 'بربري هيرو' مع 'بربري لندن'
    """
    if not isinstance(text, str): return ""
    n = text.lower()
    # إزالة الماركة (عربي + إنجليزي) — كل الأشكال
    if brand:
        for b_var in [brand.lower(), normalize(brand)]:
            n = n.replace(b_var, " ")
        # إزالة المرادفات العربية لهذه الماركة تحديداً
        brand_norm = brand.lower()
        for k, v in _SYN.items():
            if v == brand_norm or v == normalize(brand):
                n = n.replace(k, " ")
    # إزالة حروف الجر المتبقية
    for prep in ['من','في','لل','ال']:
        n = re.sub(r'\b' + prep + r'\b', ' ', n)
    # إزالة الكلمات الشائعة
    _STOP = [
        'عطر','تستر','تيستر','tester','perfume','fragrance',
        'او دو','او دي','أو دو','أو دي',
        'بارفان','بارفيوم','برفيوم','بيرفيوم','برفان','parfum','edp','eau de parfum',
        'تواليت','toilette','edt','eau de toilette',
        'كولون','cologne','edc','eau de cologne',
        'انتنس','انتينس','intense','اكستريم','extreme',
        'ابسولو','ابسوليو','absolue','absolute','absolu',
        'اكستريت','اكسترايت','extrait','extract',
        'دو','de','du','la','le','les','the',
        # أسماء ماركات فرعية تبقى بعد إزالة الماركة الرئيسية
        'هيريرا', 'تيرينزي', 'تيزيانا', 'كارولينا',
        'تيرينزي','ترينزي','terenzi','terenzio',  # Tiziana Terenzi
        'كوركدجيان','كركدجيان','kurkdjian',  # MFK
        'ميزون','مايزون','maison',  # Maison Margiela/MFK
        'باريس','paris',  # كلمة شائعة
        'دوف','dove',  # Roja Dove
        'للرجال','للنساء','رجالي','نسائي','للجنسين',
        'for men','for women','unisex','pour homme','pour femme',
        'ml','مل','ملي','milliliter',
        'كرتون ابيض','كرتون أبيض','white box',
        'اصلي','original','authentic','جديد','new',
        'اصدار','اصدارات','edition','limited',
        # كلمات شائعة ترفع pl_score خطأً
        'برفان','spray','بخاخ','عطور',
        'الرجالي','النسائي','رجال','نساء',
        'men','women','homme','femme',
        'مان','man','uomo','donna',
        'هوم','فيم',
        'او','ou','or','و',
        # كلمات إضافية ترفع pl_score خطأً
        'لو','لا','lo',
        'di','دي',
        # أجزاء أسماء الماركات المركبة التي تبقى بعد إزالة المرادف
        'جان','بول','jean','paul','gaultier',
        'كارولينا','هيريرا','carolina','herrera',
        'دولشي','غابانا','dolce','gabbana',
        'رالف','لورين','ralph','lauren',
        'ايزي','مياكي','issey','miyake',
        'فان','كليف','van','cleef','arpels',
        'اورمند','جايان','ormonde','jayne',
        'توماس','كوسمالا','thomas','kosmala',
        'فرانسيس','francis',
        'روسيندو','ماتيو','rosendo','mateu',
        'نيكولاي','nicolai',
        'ارماف','armaf',
    ]
    # إزالة الكلمات الطويلة (4+ حروف) بـ replace عادي
    # والكلمات القصيرة (1-3 حروف) بـ word boundary لمنع حذف أجزاء من كلمات أخرى
    for w in _STOP:
        if len(w) <= 3:
            n = re.sub(r'(?:^|\s)' + re.escape(w) + r'(?:\s|$)', ' ', n)
        else:
            n = n.replace(w, ' ')
    # إزالة الأرقام (الحجم) + مل/ml الملتصقة
    n = re.sub(r'\d+(?:\.\d+)?\s*(?:ml|مل|ملي)?', ' ', n)
    # إزالة الرموز
    n = re.sub(r'[^\w\s\u0600-\u06FF]', ' ', n)
    # توحيد الهمزات
    for k, v in {'أ':'ا','إ':'ا','آ':'ا','ة':'ه','ى':'ي','ؤ':'و','ئ':'ي'}.items():
        n = n.replace(k, v)
    return re.sub(r'\s+', ' ', n).strip()

def is_sample(t):
    return isinstance(t, str) and any(k in t.lower() for k in REJECT_KEYWORDS)

def is_tester(t):
    return isinstance(t, str) and any(k in t.lower() for k in TESTER_KEYWORDS)

def is_set(t):
    return isinstance(t, str) and any(k in t.lower() for k in SET_KEYWORDS)

def classify_product(name):
    """تصنيف المنتج حسب أولوية: rejected > set > tester > after_shave > body_lotion > shower_gel > hair_mist > body_mist > other > retail"""
    if not isinstance(name, str): return "retail"
    nl = name.lower()
    # ⚡ الفحص يعمل على النص الخام + النص المُطبّع معاً (يلتقط صيغ الهمزات/المرادفات)
    try:
        _nn = normalize(name)
    except Exception:
        _nn = ""
    if _nn:
        nl = (nl + " " + _nn).strip()
    # 1. مرفوض (أعلى أولوية) — عينات ومنتجات صغيرة جداً
    if any(w in nl for w in ['sample','عينة','عينه','miniature','مينياتشر','travel size','decant','تقسيم','تقسيمة','split']):
        return 'rejected'
    # ─── الجدار الفئوي — أجهزة التجميل وغير العطور ───────────────────────
    if re.search(
        r'استشوار|استشواره|مكواة|مكواه|سترتنر|straightener|dryer|hair\s*dryer'
        r'|ماسكرا|mascara|ايلاينر|eyeliner|ظل\s*عيون|eyeshadow|ايشادو'
        r'|روج|أحمر\s*شفاه|lipstick|lip\s*gloss'
        r'|بلاشر|blush|كونتور|contour|فونديشن|foundation'
        r'|فرشاة|makeup\s*brush|مرطب\s*شفاه|lip\s*balm'
        r'|طلاء\s*اظافر|nail\s*polish|nail\s*color',
        nl
    ):
        return 'other'
    # 2. طقم/مجموعة
    if any(w in nl for w in ['set ','سيت','مجموعة','gift','هدية','طقم','coffret']):
        return 'set'
    # 3. تستر
    if any(w in nl for w in ['tester','تستر','تيستر']):
        return 'tester'
    # 4. أفتر شيف
    if re.search(r'\bafter\s*shave\b|افتر\s*شيف|أفتر\s*شيف|بعد\s*الحلاقة|aftershave', nl):
        return 'after_shave'
    # 5. لوشن جسم
    if re.search(r'\bbody\s*lotion\b|لوشن\s*جسم|كريم\s*جسم|بودي\s*لوشن|body\s*cream|body\s*milk', nl):
        return 'body_lotion'
    # 6. جل استحمام / شامبو / صابون سائل
    if re.search(r'\bshower\s*gel\b|جل\s*استحمام|شاور\s*جل|body\s*wash|غسول\s*جسم|شامبو|shampoo|بلسم\b|conditioner', nl):
        return 'shower_gel'
    # 7. hair mist: كلمات كاملة فقط (لتجنب "هيريرا" → hair_mist)
    if re.search(r'\bhair\s*mist\b|عطر\s*شعر|معطر\s*شعر|بخاخ\s*شعر|للشعر|\bhair\b', nl):
        return 'hair_mist'
    # 8. body mist / سبلاش / معطر جسم: كلمات كاملة فقط
    if re.search(r'\bbody\s*mist\b|بودي\s*مست|بخاخ\s*جسم|معطر\s*جسم|سبلاش|\bsplash\b|\bbody\s*splash\b|\bbody\s*spray\b', nl):
        return 'body_mist'
    # 9. مزيل عرق/ديودرنت (v31.11c)
    if re.search(r'مزيل\s*عرق|مزيل\s*العرق|ديودرنت|ديودورنت|deodorant|deo\b|anti.?perspirant|مضاد\s*التعرق', nl):
        return 'deodorant'
    # 10. صابون (v31.11c)
    if re.search(r'\bsoap\b|صابون|صابونة|صابونه|bar\s*soap', nl):
        return 'soap'
    # 11. بودرة/كريم/لوشن عام
    if re.search(r'بودرة|بودره|powder|كريم|cream|لوشن|lotion', nl):
        return 'other'
    return 'retail'

def _price(row):
    for c in ["السعر", "سعر المنتج", "سعر_المنتج", "Price", "price", "سعر", "PRICE", "السعر بعد الخصم"]:
        if c in row.index:
            try: return float(str(row[c]).replace(",",""))
            except (ValueError, TypeError): pass
    # احتياطي مُقيَّد: فقط أعمدة تبدو سعرية (يمنع خلط SKU/كمية بالسعر)
    for c in row.index:
        cl = str(c).lower()
        if "سعر" not in cl and "price" not in cl and "cost" not in cl and "ثمن" not in cl:
            continue  # ← تخطي الأعمدة غير السعرية
        try:
            v = float(str(row[c]).replace(",",""))
            if 1 <= v <= 99999:  # نطاق سعر معقول
                return v
        except (ValueError, TypeError):
            pass
    return 0.0

def _pid(row, id_col):
    """استخراج معرف المنتج من الصف — آمن للقيم الفارغة."""
    if not id_col or id_col not in row.index:
        return ""
    v = row.get(id_col, "")
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    s = str(v).strip()
    if s.lower() in ("nan", "none", "<na>", ""):
        return ""
    try:
        fv = float(s.replace(",", ""))
        return str(int(fv)) if fv == int(fv) else s
    except (ValueError, TypeError):
        return s

def _fcol(df, cands):
    """بحث مرن عن العمود — يعيد None إذا لم يجد تطابقاً."""
    cols = list(df.columns)
    # بحث 1: تطابق تام
    for c in cands:
        if c in cols: return c
    # بحث 2: تطبيع الهمزات (أ/إ/آ → ا)
    def _norm_ar(s):
        return str(s).replace('أ','ا').replace('إ','ا').replace('آ','ا').strip()
    norm_cols = {_norm_ar(c): c for c in cols}
    for c in cands:
        nc = _norm_ar(c)
        if nc in norm_cols: return norm_cols[nc]
    # بحث 3: بحث جزئي (العمود يحتوي على الكلمة المفتاحية)
    for c in cands:
        for col in cols:
            if c in col or _norm_ar(c) in _norm_ar(col):
                return col
    return None


def _fcol_optional(df, cands):
    """مثل _fcol لكن بدون الرجوع للعمود الأول عند عدم التطابق — يعيد None."""
    if df is None or df.empty:
        return None
    cols = list(df.columns)
    for c in cands:
        if c and c in cols:
            return c

    def _norm_ar(s):
        return str(s).replace("أ", "ا").replace("إ", "ا").replace("آ", "ا").strip()

    norm_map = {_norm_ar(str(c)): c for c in cols}
    for c in cands:
        if not c:
            continue
        nc = _norm_ar(c)
        if nc in norm_map:
            return norm_map[nc]
    for c in cands:
        if not c:
            continue
        for col in cols:
            if c in str(col) or _norm_ar(c) in _norm_ar(str(col)):
                return col
    return None


def _find_image_column(df):
    """عمود صورة المنتج — يشمل تصدير سلة ([n] الصورة / اللون) ومرادفات."""
    if df is None or df.empty:
        return None
    c = _fcol_optional(df, [
        "صورة المنتج", "صوره المنتج", "image", "Image", "product_image", "الصورة",
        "الصورة / اللون", "[1] الصورة / اللون", "[2] الصورة / اللون", "[3] الصورة / اللون",
        "thumbnail", "Thumb", "photo",
    ])
    if c:
        return c
    for col in df.columns:
        sc = str(col)
        if "وصف صورة" in sc or "وصف صوره" in sc:
            continue
        if "صورة" in sc or "image" in sc.lower():
            return col
        if "thumb" in sc.lower() and "url" not in sc.lower():
            return col
    return None


def _find_url_column(df):
    """عمود رابط صفحة المنتج — يشمل أعمدة CSV الغريبة مثل abs-size href."""
    if df is None or df.empty:
        return None
    c = _fcol_optional(df, [
        "رابط المنتج", "الرابط", "رابط", "product_url", "Product URL",
        "link", "url", "URL", "product link",
    ])
    if c:
        return c
    for col in df.columns:
        sc = str(col)
        sl = sc.lower()
        if "صورة" in sc and "وصف" not in sc and "رابط" not in sc:
            continue
        if "href" in sl or "رابط" in sc or ("link" in sl and "image" not in sl):
            return col
        if sl in ("url", "uri") or sc.endswith(" URL"):
            return col
    return None


def _header_looks_like_price_or_link(col) -> bool:
    """رؤوس لا تُستخدم كعمود «اسم المنتج» بالخطأ (مثل رابط المنتج / سعر المنتج)."""
    sc = str(col).strip()
    sl = sc.lower()
    if "رابط" in sc:
        return True
    if "سعر" in sc and "وصف" not in sc:
        return True
    if "href" in sl or sl in ("url", "uri") or sl.endswith(" url"):
        return True
    if "price" in sl and "repr" not in sl:
        return True
    return False


def _find_product_name_column(df):
    """
    عمود اسم المنتج — يتجنب مطابقة «المنتج» الجزئية داخل «رابط المنتج» أو «سعر المنتج»
    (انظر _fcol: كان يعيد رابط المنتج كاسم عند ملف عمودين).
    """
    if df is None or df.empty:
        return ""

    def _norm(s):
        return str(s).replace("أ", "ا").replace("إ", "ا").replace("آ", "ا").strip()

    cols = list(df.columns)

    # 1) تطابق تام
    for cand in (
        "اسم المنتج", "المنتج", "Product", "Name", "name",
        "Title", "title", "الاسم", "Product Name", "product name",
    ):
        if cand in cols and not _header_looks_like_price_or_link(cand):
            return cand
    for col in cols:
        nc = _norm(str(col))
        for cand in (
            "اسم المنتج", "المنتج", "Product", "Name", "name", "Title", "title",
        ):
            if _norm(cand) == nc and not _header_looks_like_price_or_link(col):
                return col

    # 2) تطابق جزئي — تخطّ الرؤوس التي تبدو رابطاً أو سعراً
    for cand in (
        "اسم المنتج", "المنتج", "Product", "Name", "name", "Title", "title",
    ):
        for col in cols:
            if _header_looks_like_price_or_link(col):
                continue
            sc = str(col)
            if cand in sc or _norm(cand) in _norm(sc):
                return col

    # 3) أول عمود ليس رابطاً معروفاً ولا عمود سعراً معروفاً
    url_c = _find_url_column(df)
    price_c = _fcol_optional(df, [
        "سعر المنتج", "السعر", "سعر", "Price", "price", "PRICE",
    ])
    for col in cols:
        if col == url_c or col == price_c:
            continue
        if _header_looks_like_price_or_link(col):
            continue
        return col

    # 4) عمودان (رابط + سعر): لا تستخدم عمود الرابط كاسم — خذ الآخر (غالباً السعر؛ يُفضّل ضبط الدور يدوياً)
    if url_c and len(cols) == 2:
        for col in cols:
            if col != url_c:
                return col

    return cols[0] if cols else ""


def _name_col_for_analysis(df):
    """بعد apply_user_column_map: «المنتج» القياسي أو تعرف آمن لاسم المنتج."""
    if df is None or df.empty:
        return ""
    if "المنتج" in df.columns:
        return "المنتج"
    return _find_product_name_column(df)


def _first_product_page_url_from_row(row):
    """أول رابط http لا يبدو ملف صورة مباشر (صفحة منتج)."""
    for c in row.index:
        v = row.get(c)
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        s = str(v).strip()
        if not s.startswith("http"):
            continue
        if _looks_like_image_url(s):
            continue
        return s.split()[0]
    return ""


def _first_image_url_from_row(row):
    """أول قيمة في الصف تبدو رابط صورة (عندما تكون الخلية تحت عمود خاطئ أو فارغ)."""
    for c in row.index:
        u = _extract_image_url_from_cell(row.get(c))
        if u:
            return u
    return ""


def resolve_catalog_columns(df):
    """أسماء أعمدة متجرنا بعد القراءة — للكتالوج وواجهة vs_card."""
    if df is None or df.empty:
        return {"name": "", "price": "", "id": "", "img": "", "url": ""}
    return {
        "name": _find_product_name_column(df),
        "price": _fcol(df, ["سعر المنتج", "السعر", "سعر", "Price", "price", "PRICE"]),
        "id": _fcol(df, [
            "رقم المنتج", "معرف المنتج", "المعرف", "معرف", "رقم_المنتج", "معرف_المنتج",
            "product_id", "Product ID", "Product_ID", "ID", "id", "Id",
            "SKU", "sku", "Sku", "رمز المنتج", "رمز_المنتج", "رمز المنتج sku",
            "الكود", "كود", "Code", "code", "الرقم", "رقم", "Barcode", "barcode", "الباركود",
        ]),
        "img": (_find_image_column(df) or ""),
        "url": (_find_url_column(df) or ""),
    }


def detect_input_columns(df):
    """
    تعرف تلقائي على أعمدة ملف المتجر أو ملف المنافس (بعد read_file).
    يُرجع أسماء الأعمدة المربوطة: اسم، سعر، معرف، صورة، رابط + قائمة كل الأعمدة.
    """
    if df is None or df.empty:
        return {"ok": False, "error": "ملف فارغ أو غير مقروء"}
    m = resolve_catalog_columns(df)
    nc = len(df.columns)
    return {
        "ok": True,
        "columns_count": nc,
        "role_hint": "ملف_منافس_محتمل" if nc <= 8 else "ملف_متجر_أو_تصدير_واسع",
        "mapping": {
            "اسم المنتج ← عمود": m.get("name") or "—",
            "سعر المنتج ← عمود": m.get("price") or "—",
            "معرف / SKU ← عمود": m.get("id") or "—",
            "صورة المنتج ← عمود": m.get("img") or "—",
            "رابط المنتج ← عمود": m.get("url") or "—",
        },
        "all_column_names": [str(c) for c in df.columns],
    }


# أسماء قياسية يبحث عنها المحرك عبر _fcol / _find_*_column
_USER_MAP_CANON = {
    "name": "المنتج",
    "price": "سعر المنتج",
    "id": "رقم المنتج",
    "img": "صورة المنتج",
    "url": "رابط المنتج",
}


def apply_user_column_map(df, name=None, price=None, id_col=None, img=None, url=None):
    """
    يطبّق اختيار المستخدم من القوائم المنسدلة: ينسخ كل عمود مختار إلى عمود قياسي
    (المنتج، سعر المنتج، …) ليتعرّف عليه المحرك دون كسر بقية الأعمدة.
    """
    if df is None or df.empty:
        return df
    out = df.copy()
    pairs = [
        ("name", name),
        ("price", price),
        ("id", id_col),
        ("img", img),
        ("url", url),
    ]
    skip = frozenset(("", "—", "— (تخطي)", "nan", "None", None))
    for role, src in pairs:
        if src is None or str(src).strip() in skip:
            continue
        sc = str(src).strip()
        if sc not in out.columns:
            continue
        canon = _USER_MAP_CANON[role]
        try:
            out[canon] = out[sc]
        except Exception:
            out[canon] = out[sc].astype(str)
    return out


# ═══════════════════════════════════════════════════════
# ⚡ v31.8: ثوابت مسبقة الحساب — خارج الحلقات
# ═══════════════════════════════════════════════════════
_NUM_WORDS = {
    'ون':'1','تو':'2','ثري':'3','فور':'4','فايف':'5',
    'سكس':'6','سفن':'7','ايت':'8','ناين':'9','تن':'10',
    'one':'1','two':'2','three':'3','four':'4','five':'5',
    'six':'6','seven':'7','eight':'8','nine':'9','ten':'10',
    'i':'1','ii':'2','iii':'3','iv':'4','v':'5',
    'vi':'6','vii':'7','viii':'8','ix':'9','x':'10',
}

_RE_PROD_NUM  = re.compile(r'(?:no|num|number|نمبر|رقم|№|#)\s*(\d+)', re.IGNORECASE)
_RE_STUCK_NUM = re.compile(r'[a-z\u0600-\u06FF](\d+)', re.IGNORECASE)
_RE_SOLO_NUM  = re.compile(r'\b(\d{1,3})\b')
_SIZE_NUMBERS = {'100','50','30','200','150','75','80','125','250','300','ml'}
_PROD_NUMBERS = {'212','360','1','2','3','4','5','6','7','8','9','11','12','13','14','15','16','17','18','19','21'}

def _extract_product_numbers(text):
    """Extract product-identifying numbers (not sizes) — module-level for speed."""
    nums = set()
    tl = text.lower()
    for m in _RE_PROD_NUM.finditer(tl):
        nums.add(m.group(1))
    for word, num in _NUM_WORDS.items():
        if f'نمبر {word}' in tl or f'number {word}' in tl or f'no {word}' in tl or f'رقم {word}' in tl:
            nums.add(num)
    for m in _RE_STUCK_NUM.finditer(tl):
        v = m.group(1)
        if v not in _SIZE_NUMBERS:
            nums.add(v)
    for m in _RE_SOLO_NUM.finditer(tl):
        v = m.group(1)
        pos = m.end()
        after = tl[pos:pos+5].strip()
        if after.startswith('ml') or after.startswith('مل'):
            continue
        if v in _PROD_NUMBERS:
            nums.add(v)
    return nums

# v31.10d: مجموعات مترادفة — كل متغيرات نفس المفهوم في مجموعة واحدة
# "oud"/"عود"/"العود" = مجموعة واحدة → لا رفض خاطئ عند اختلاف اللغة
_FLANKER_GROUPS = [
    ('sport', 'سبورت'),
    ('intense', 'انتنس', 'إنتنس', 'انتينس', 'إنتينس'),
    ('elixir', 'الكسير', 'إلكسير', 'اليكسير', 'إليكسير', 'اليكسر'),
    ('oud', 'عود', 'العود'),
    ('absolu', 'ابسولو', 'absolue', 'ابسوليت', 'ابسولوت', 'absolute', 'ابسوليو', 'ابسلو', 'أبسلو', 'أبسول', 'ابسول', 'ابسلوتلي', 'أبسولوتلي', 'absolutely'),
    ('leather', 'ليذر'),
    ('black', 'بلاك'),
    ('extreme', 'اكستريم', 'إكستريم'),
    ('poudree', 'بودريه'),
    ('flame', 'فليم'),
    ('night', 'نايت'),
    ('gold', 'جولد', 'غولد', 'قولد'),
    ('aqua', 'أكوا', 'اكوا'),
    ('nuit', 'نوي'),
    ('blanc', 'بلانك'),
    ('prive', 'privé', 'بريفيه'),
    ('legend', 'ليجند'),
    ('royal', 'رويال'),
    ('wild', 'وايلد', 'وايد'),
    ('encens', 'اينسينس', 'انسينس'),
    ('nectar', 'نكتار'),
    ('aura', 'اورا'),
    ('idol', 'ايدول', 'آيدول'),
    ('now', 'ناو'),
    ('grand', 'جراند'),
    ('tuberose', 'توبيروز', 'توبيروزا', 'تيوبروزا', 'تيوبيروزا'),
    ('saffron', 'سافرون'),
    ('amethyst', 'اميثيست'),
    ('irish', 'ايرش', 'ايريش', 'ايرس', 'ايريس', 'ايرز', 'آيريس', 'آيرز'),
    ('burning', 'بورنينج', 'بيرنينج', 'بيرننج'),
    ('confidential', 'كونفيدنشال'),
    ('atlantide', 'اتلانتيدى', 'اتلانتيدي'),
    ('kirke', 'كيركي'),
    ('andromeda', 'اندروميدا'),
    ('chiron', 'شيرون'),
    ('tilia', 'تيليا'),
    ('orion', 'اوريون'),
    ('cassiopea', 'كاسيوبيا'),
    ('anthology', 'انثولوجي'),
    ('halfeti', 'هالفيتي'),
    ('sartorial', 'سارتوريال'),
    ('babylon', 'بابيلون'),
    ('lazuli', 'لاوزلي', 'لازولي', 'لازوري'),
    ('empire', 'امباير', 'إمباير'),
    ('mademoiselle', 'مادموزيل', 'مادموازيل', 'مودموزيل', 'مادمازل', 'مدموزيل'),
    ('noir', 'نوار', 'نوير'),
    ('premiere', 'بريميير'),
    ('tenderness', 'تندرنس'),
    ('rose', 'روز'),
    ('silver', 'سيلفر', 'سلفر'),
    ('crystal', 'كريستال'),
    ('passione', 'باسيوني', 'باسيون'),
    ('floral', 'فلورال'),
    ('bloom', 'بلوم'),
    ('ambrosia', 'امبروسيا'),
    ('allure', 'الور', 'ألور'),
    ('donna', 'دونا'),
    ('uomo', 'اومو'),
    ('born', 'بورن'),
    ('phantom', 'فانتوم'),
    ('scandal', 'سكاندال'),
    ('desire', 'ديزاير'),
    ('essence', 'ايسنس', 'إيسنس', 'اسينس'),
    ('exclusif', 'اكسكلوسيف', 'إكسكلوسيف', 'اكسكلوسيفس', 'إكسكلوسيفس'),
    ('fresh', 'فريش'),
    ('luminous', 'لومينوس'),
    ('stars', 'ستارز'),
    ('walker', 'ووكر'),
    ('momentum', 'مومينتوم'),
    ('queen', 'كوين'),
    ('petite', 'بوتيت'),
    ('perfecto', 'بيرفيكتو'),
    ('xs',),
    ('braun', 'براون'),
    ('rockstuded', 'روكستد'),
    ('المسك',),
]

_FLANKER_GROUP_RES = []
for group in _FLANKER_GROUPS:
    alt = '|'.join(re.escape(w) for w in group)
    try:
        pat = re.compile(r'(?:\b|(?<=\s)|(?<=^))(?:' + alt + r')(?:\b|(?=\s)|(?=$))', re.IGNORECASE)
    except re.error:
        pat = None
    _FLANKER_GROUP_RES.append(pat)

def _group_in_text(group_idx, text):
    """Check if any variant of the flanker group is found in text."""
    pat = _FLANKER_GROUP_RES[group_idx]
    if pat and pat.search(text):
        return True
    for variant in _FLANKER_GROUPS[group_idx]:
        if variant in text:
            return True
    return False


# ═══════════════════════════════════════════════════════
#  الكلاس الجديد: Pre-normalized Competitor Index
#  يُبنى مرة واحدة لكل ملف منافس ← يسرّع الـ matching 5x
# ═══════════════════════════════════════════════════════
class CompIndex:
    """⚡ v31.8: فهرس المنافس — يستخدم الاستخراج المسبق من DB إذا توفر"""
    def __init__(self, df, name_col, id_col, comp_name, img_col=None, url_col=None):
        self.comp_name = comp_name
        self.name_col  = name_col
        self.id_col    = id_col
        self.img_col   = (img_col or "") or ""
        self.url_col   = (url_col or "") or ""
        self.df        = df.reset_index(drop=True)
        self.raw_names  = self.df[self.name_col].fillna("").astype(str).tolist()
        self.norm_names = [normalize(n) for n in self.raw_names]

        # ⚡ v31.8: استخدم الأعمدة المسبقة إذا توفرت (من DB) — أسرع 30x
        _has_pre = (
            "agg_name" in self.df.columns
            and self.df["agg_name"].notna().any()
            and (self.df["agg_name"].fillna("").astype(str) != "").sum() > len(self.df) * 0.5
        )

        if _has_pre:
            # DB pre-computed — تحميل فوري
            self.agg_names = self.df["agg_name"].fillna("").astype(str).tolist()
            self.brands    = self.df["extracted_brand"].fillna("").astype(str).tolist()
            self.sizes     = pd.to_numeric(self.df["extracted_size"], errors='coerce').fillna(0).tolist()
            # v34: التركيز يُعاد حسابه طازجاً دائماً (عمود extracted_type قديم بمنطق ضعيف)
            self.types     = [extract_type(n) for n in self.raw_names]
            self.genders   = self.df["extracted_gender"].fillna("").astype(str).tolist()
            self.plines    = self.df["product_line"].fillna("").astype(str).tolist()
            self.classes   = self.df["extracted_class"].fillna("").astype(str).tolist()

            # v22: استخدام عمود brand الأصلي (88% تغطية) قبل extracted_brand (44%)
            _raw_brand_col = None
            for _bc in ("brand", "الماركة"):
                if _bc in self.df.columns:
                    _raw_brand_col = _bc
                    break
            _raw_brands = (
                self.df[_raw_brand_col].fillna("").astype(str).str.strip().tolist()
                if _raw_brand_col else [""] * len(self.df)
            )

            # ملء القيم الفارغة: brand column → extracted_brand → extract_brand(name)
            for i, n in enumerate(self.raw_names):
                if not self.brands[i]:
                    # Try raw brand column first (from store/scraper)
                    if _raw_brands[i]:
                        self.brands[i] = extract_brand(_raw_brands[i]) or _raw_brands[i]
                    else:
                        self.brands[i] = extract_brand(n) or ""
                if not self.agg_names[i]:
                    self.agg_names[i] = normalize_name(n) or ""
                if not self.classes[i]:
                    self.classes[i] = classify_product(n) or ""
        else:
            # لا يوجد استخراج مسبق — حساب كامل (ملفات مرفوعة يدوياً)
            # v22: prefer brand/الماركة column if available
            _raw_brand_col2 = None
            for _bc2 in ("brand", "الماركة"):
                if _bc2 in self.df.columns:
                    _raw_brand_col2 = _bc2
                    break
            _raw_brands2 = (
                self.df[_raw_brand_col2].fillna("").astype(str).str.strip().tolist()
                if _raw_brand_col2 else [""] * len(self.df)
            )
            self.agg_names  = [normalize_name(n) for n in self.raw_names]
            self.brands     = []
            for i, n in enumerate(self.raw_names):
                br = extract_brand(n) or ""
                if not br and _raw_brands2[i]:
                    br = extract_brand(_raw_brands2[i]) or _raw_brands2[i]
                self.brands.append(br)
            self.sizes      = [extract_size(n) for n in self.raw_names]
            self.types      = [extract_type(n) for n in self.raw_names]
            self.genders    = [extract_gender(n) for n in self.raw_names]
            self.plines     = [extract_product_line(n, self.brands[i]) for i, n in enumerate(self.raw_names)]
            self.classes    = [classify_product(n) for n in self.raw_names]

        # ⚡ v31.8: pre-compute prices بدل iterrows البطيء
        _price_cols = [c for c in ['السعر','سعر المنتج','Price','price'] if c in self.df.columns]
        if _price_cols:
            self.prices = pd.to_numeric(self.df[_price_cols[0]].astype(str).str.replace(',','',regex=False), errors='coerce').fillna(0).tolist()
        else:
            self.prices = [_price(row) for _, row in self.df.iterrows()]
        _id_cols = [id_col] if id_col and id_col in self.df.columns else []
        if _id_cols:
            self.ids = self.df[_id_cols[0]].fillna('').astype(str).str.strip().tolist()
        else:
            self.ids = [''] * len(self.df)
        n = len(self.df)
        if self.img_col and self.img_col in self.df.columns:
            self.extra_imgs = self.df[self.img_col].fillna("").astype(str).str.strip().tolist()
        else:
            self.extra_imgs = [""] * n
        if self.url_col and self.url_col in self.df.columns:
            self.extra_urls = self.df[self.url_col].fillna("").astype(str).str.strip().tolist()
        else:
            self.extra_urls = [""] * n

        # ⚡ v31.8: pre-compute normalized brands + product numbers
        self.norm_brands = [normalize(b).lower() if b else '' for b in self.brands]
        self.prod_nums   = [_extract_product_numbers(n) for n in self.norm_names]

        # ⚡ v33: Brand Index
        self._brand_index: dict[str, list[int]] = {}
        for i, nbr in enumerate(self.norm_brands):
            if nbr not in self._brand_index:
                self._brand_index[nbr] = []
            self._brand_index[nbr].append(i)

        # ⚡ v22: pre-compute non-sample set ONCE (was 108K×7928 calls)
        self._nonsample_set = frozenset(i for i, n in enumerate(self.raw_names) if not is_sample(n))

    def search(self, our_norm, our_br, our_sz, our_tp, our_gd, our_pline="", top_n=6, our_price=0):
        """⚡ v31.8: بحث محسّن — ثوابت خارج الحلقة + pre-compiled regex"""
        if not self.norm_names: return []

        # ⚡ v31.8: normalize(our_br) مرة واحدة فقط
        _our_br_norm = normalize(our_br).lower() if our_br else ""

        # ⚡ v22: Brand-First Search is MANDATORY (matches mandatory brand filter)
        if not _our_br_norm:
            return []

        _brand_candidates = self._brand_index.get(_our_br_norm, [])
        if not _brand_candidates:
            return []  # No competitors with this brand

        # intersect brand candidates with pre-computed non-sample set
        valid_idx = [i for i in _brand_candidates if i in self._nonsample_set]
        if not valid_idx:
            return []

        valid_aggs = [self.agg_names[i] for i in valid_idx]

        our_agg = normalize_name(our_norm) if our_norm else our_norm
        fast = rf_process.extract(
            our_agg, valid_aggs,
            scorer=fuzz.token_set_ratio,
            limit=min(30, len(valid_aggs))
        )

        cands = []
        seen  = set()
        our_class = classify_product(our_norm)
        # ⚡ v31.8: pre-compute per-search constants
        our_pnums = _extract_product_numbers(our_norm) if our_norm else set()
        o_n_low = our_norm.lower() if our_norm else ""

        for _, fast_score, vi in fast:
            if fast_score < 45: continue
            idx  = valid_idx[vi]
            name = self.raw_names[idx]
            if name in seen: continue

            c_br = self.brands[idx]
            c_sz = self.sizes[idx]
            c_tp = self.types[idx]
            c_gd = self.genders[idx]
            c_pl = self.plines[idx]

            # ═══ فلاتر سريعة ═══
            # v22: brand filter is MANDATORY — no match without confirmed brand
            if _our_br_norm and c_br:
                if _our_br_norm != self.norm_brands[idx]: continue
            else:
                # One or both brands unknown → cannot confirm match → reject
                continue
            if our_sz > 0 and c_sz > 0 and abs(our_sz - c_sz) > 2: continue
            # v22: reject size/no-size mismatch — one has size, other doesn't
            if (our_sz > 0 and c_sz == 0) or (our_sz == 0 and c_sz > 0): continue
            # ═══ v34: التركيز فلتر صارم — اختلاف التركيز المعروف = SKU مختلف = رفض ═══
            #  (شانيل بلو اكسكلوسيف ≠ EDP ≠ EDT ≠ Parfum ≠ Elixir).
            #  النسخة المكثّفة (+INT) تُعامَل كتركيز مختلف عن العادي.
            if our_tp and c_tp and our_tp != c_tp:
                continue
            if our_gd and c_gd and our_gd != c_gd: continue

            # ═══ فلتر تصنيف المنتج (v34: صارم — لا نطابق العطر بمنتج غير عطري) ═══
            #  مزيل عرق / لوشن بعد الحلاقة / جل استحمام / صابون / بخاخ شعر = ليست عطراً → رفض.
            c_class = self.classes[idx]
            _NONPERF = ('rejected','hair_mist','body_mist','set','other',
                        'after_shave','deodorant','body_lotion','shower_gel','soap')
            if our_class != c_class:
                if our_class in _NONPERF or c_class in _NONPERF:
                    continue
                if (our_class == 'tester') != (c_class == 'tester'):
                    continue
            # v31.11: حارس المجموعات/الطقم — لا نطابق مجموعة مع عطر فردي أبداً
            _our_is_set = any(w in o_n_low for w in ('مجموعة','مجموعه','طقم','gift set','gift box'))
            _comp_is_set = any(w in (name.lower() if name else '') for w in ('مجموعة','مجموعه','طقم','gift set','gift box'))
            if _our_is_set != _comp_is_set:
                continue
            # حتى لو تطابق التصنيف، احرس أسماء المنتجات غير العطرية صراحةً في اسم المنافس
            _cl = name.lower()
            if any(w in _cl for w in (
                    # مزيلات العرق — كل الصيغ
                    "مزيل للرائحة","مزيل العرق","مزيل رائحة","مزيل عرق",
                    "مزيل للعرق","ديودرنت","ديودرانت","deodorant",
                    "anti-perspirant","antiperspirant","رول اون","roll on",
                    "رذاذ مزيل","بخاخ مزيل","سبراي مزيل",
                    # بعد الحلاقة
                    "after shave","aftershave","بعد الحلاقة","افتر شيف",
                    # لوشن وكريم
                    "لوشن","لوسيون","بودي لوشن","body lotion",
                    "كريم جسم","body cream","كريم يد","hand cream",
                    # جل استحمام
                    "جل استحمام","شاور جل","shower gel","body wash","غسول جسم",
                    # صابون وشامبو
                    "صابون","شامبو","shampoo","soap",
                    # بخاخ جسم
                    "بخاخ جسم","بخاخ للجسم","بودي سبراي","body spray",
                    "body mist","معطر جسم","معطر للجسم","بودي مست",
                    "رذاذ مرطب","رذاذ جسم","رذاذ للجسم",
                    # عطر شعر / زيت شعر وجسم
                    "عطر شعر","معطر شعر","hair mist","hair perfume",
                    "زيت شعر","زيت الشعر","hair oil",
                    "زيت الشعر والجسم","زيت جسم","body oil",
                    )):
                # منتجنا عطر (retail/tester) لكن المنافس منتج عناية → ليس نفس المنتج
                if our_class in ('retail','tester'):
                    continue

            # ═══ مقارنة الأرقام — pre-computed ═══
            c_pnums = self.prod_nums[idx]
            if our_pnums and c_pnums and our_pnums != c_pnums:
                continue

            # ═══ مقارنة خط الإنتاج (v31.11b: token_sort + length check) ═══
            # token_sort_ratio بدلاً من token_set_ratio لأن الأخير يعطي 100%
            # عندما يكون اسم قصير جزء من اسم طويل (مثل "بيل" ⊂ "بيل بارادايس جاردن")
            pline_penalty = 0
            if our_pline and c_pl:
                pl_score = fuzz.token_sort_ratio(our_pline, c_pl)
                # عقوبة إضافية: إذا أحد الخطين أطول بكثير من الآخر
                # (مثل "بيل" vs "بيل بارادايس جاردن" = نسبة طول 20%)
                _pl_words_our = set(our_pline.split())
                _pl_words_comp = set(c_pl.split())
                _pl_len_ratio = min(len(_pl_words_our), len(_pl_words_comp)) / max(len(_pl_words_our), len(_pl_words_comp), 1)
                if our_br and c_br:
                    # v31.11b: رفض إذا pl_score < 60 أو اختلاف طول كبير مع score < 70
                    if pl_score < 60:
                        continue
                    if pl_score < 70 and _pl_len_ratio < 0.5:
                        continue  # "بيل" vs "بيل بارادايس جاردن" → رفض
                    elif pl_score < 75:
                        pline_penalty = -35
                    elif pl_score < 88:
                        pline_penalty = -15
                else:
                    if pl_score < 50:
                        continue
                    elif pl_score < 65:
                        pline_penalty = -45
                    elif pl_score < 80:
                        pline_penalty = -25

            # ═══ score تفصيلي ═══
            n1 = our_agg
            n2 = self.agg_names[idx]
            s1 = fuzz.token_sort_ratio(n1, n2)
            s2 = fuzz.token_set_ratio(n1, n2)
            s3 = fuzz.partial_ratio(n1, n2)
            base = s1*0.30 + s2*0.50 + s3*0.20

            # ═══ تعديلات الماركة — pre-computed (v22: brand already confirmed in filter) ═══
            # If we reach here, brands match (filter above guarantees it)
            base += 10

            if not our_pline or not c_pl:
                base -= 20

            # ═══ تعديلات الحجم (v22: strict 2ml tolerance) ═══
            if our_sz > 0 and c_sz > 0:
                d = abs(our_sz - c_sz)
                base += 10 if d == 0 else (-5 if d <= 2 else -100)

            # v31.10c: حارس نسبة السعر — فرق >4x يعني منتج مختلف
            if our_price and our_price > 50:
                c_price = self.prices[idx] if idx < len(self.prices) else 0
                if c_price and c_price > 0:
                    _pr = max(our_price / c_price, c_price / our_price)
                    if _pr > 4.0:
                        base -= 45

            if our_tp and c_tp and our_tp != c_tp:
                base -= 40

            if our_gd and c_gd and our_gd != c_gd:
                continue
            elif (our_gd or c_gd) and our_gd != c_gd:
                base -= 15

            base += pline_penalty

            # ═══ Flankers Guard — مجموعات مترادفة (v31.10d) ═══
            c_n_low = name.lower() if name else ""
            _flanker_mismatch = False
            for gi in range(len(_FLANKER_GROUPS)):
                in_our = _group_in_text(gi, o_n_low)
                in_comp = _group_in_text(gi, c_n_low)
                if in_our != in_comp:
                    _flanker_mismatch = True
                    break
            if _flanker_mismatch:
                if our_br and c_br:
                    continue  # رفض مباشر — نفس الماركة لكن خط منتج مختلف
                else:
                    base -= 45

            # ═══ v31.11b: حارس الكلمات المميزة الإلزامية ═══
            # إذا منتجنا يحتوي كلمة مميزة مهمة (بلومينغ، كاربون، بلانش...)
            # والمنافس لا يحتوي هذه الكلمة أو مرادفها → رفض
            if our_br and c_br:
                _MUST_MATCH_WORDS = {
                    # (our_word, synonym_set) — إذا our_word في منتجنا، يجب أحد المرادفات في المنافس
                    'بلومينغ': {'بلومينغ','بلومنغ','blooming'},
                    'كاربون': {'كاربون','carbon'},
                    'أوشن': {'أوشن','اوشن','ocean'},
                    'سبورت': {'سبورت','sport'},
                    'بلانش': {'بلانش','blanche','بلانك'},
                    'تندر': {'تندر','تاندر','تيندر','tendre','tender'},
                    'فراش': {'فراش','فريش','fraiche','فراشي'},
                    'مادموزيل': {'مادموزيل','مادموازيل','mademoiselle'},
                    'بودريه': {'بودريه','بودري','poudree','بودره'},
                    'اونيكس': {'اونيكس','أونيكس','onyx'},
                    'سيلستيا': {'سيلستيا','celestia'},
                    'بارادايس': {'بارادايس','paradise'},
                    'سنترال بارك': {'سنترال بارك','central park'},
                    'فري رايد': {'فري رايد','free ride'},
                    'بيرلز': {'بيرلز','بيرل','pearl','perle'},
                    'اكسكلوسيف': {'اكسكلوسيف','إكسكلوسيف','exclusif','exclusive','حصري'},
                    'بلاتينوم': {'بلاتينوم','platinum'},
                    'كريستال': {'كريستال','crystal','كريستالي'},
                    'نكتار': {'نكتار','nectar'},
                    # v31.11c: كلمات إضافية رُصدت من الفحص
                    'فلاور': {'فلاور','فلاورز','فلورز','فلوارز','flower','flowers','fleur','fleurs'},
                    'ليجير': {'ليجير','ليجيره','ليجيري','ليجر','legere','léger','légère','leger'},
                    'فيتيفر': {'فيتيفر','فيتفير','فتيفر','فيتفر','فيتيفير','vetiver','vétiver'},
                    'ابسولو': {'ابسولو','أبسولو','ابسوليو','ابسلوت','أبسولوت','ابسولوت','absolu','absolue','absolute'},
                }
                _dist_reject = False
                for _dw, _syns in _MUST_MATCH_WORDS.items():
                    if any(s in o_n_low for s in _syns):
                        if not any(s in c_n_low for s in _syns):
                            _dist_reject = True
                            break
                if _dist_reject:
                    continue

            score = round(max(0, min(100, base)), 1)
            if score < 60: continue

            seen.add(name)
            img_u = self.extra_imgs[idx] if idx < len(self.extra_imgs) else ""
            url_u = self.extra_urls[idx] if idx < len(self.extra_urls) else ""
            cands.append({
                "name": name, "score": score,
                "price": self.prices[idx], "product_id": self.ids[idx],
                "brand": c_br, "size": c_sz, "type": c_tp, "gender": c_gd,
                "competitor": self.comp_name,
                "image_url": img_u, "product_url": url_u,
                "thumb": img_u,
            })

        cands.sort(key=lambda x: x["score"], reverse=True)
        return cands[:top_n]


# ═══════════════════════════════════════════════════════
#  AI Batch — Gemini + OpenRouter fallback
# ═══════════════════════════════════════════════════════
_GURL    = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent"
_OR_URL  = "https://openrouter.ai/api/v1/chat/completions"
_OR_FREE = [
    "meta-llama/llama-3.3-70b-instruct:free",
    "google/gemini-2.0-flash-exp:free",
    "deepseek/deepseek-chat-v3-0324:free",
    "mistralai/mistral-7b-instruct:free",
]

# ⚡ v32: Cooldown ذكي للمفاتيح — يتخطى المفاتيح المحظورة 60 ثانية بدل الانتظار
import time as _time_mod
_KEY_COOLDOWN = {}  # {key_hash: timestamp_when_blocked}
_KEY_COOLDOWN_SEC = 60  # تجاهل مفتاح محظور لمدة 60 ثانية

def _key_available(key):
    kh = key[-8:]
    blocked_at = _KEY_COOLDOWN.get(kh, 0)
    return (_time_mod.time() - blocked_at) > _KEY_COOLDOWN_SEC

def _key_block(key):
    _KEY_COOLDOWN[key[-8:]] = _time_mod.time()

def _ai_batch(batch):
    """
    batch: [{"our":str, "price":float, "candidates":[...]}]
    → [int]  (0-based index | -1=no match)
    يحاول Gemini أولاً ثم OpenRouter تلقائياً — لا يتوقف أبداً
    """
    if not batch:
        return []

    # ── cache ────────────────────────────────────────────────────────────
    ck = hashlib.md5(json.dumps(
        [{"o": x["our"], "c": [c["name"] for c in x["candidates"]]} for x in batch],
        ensure_ascii=False, sort_keys=True).encode()).hexdigest()
    cached = _cget(ck)
    if cached is not None:
        return cached

    # ── بناء الـ prompt ───────────────────────────────────────────────────
    lines = []
    for i, it in enumerate(batch):
        cands = "\n".join(
            f"  {j+1}. {c['name']} | {int(c.get('size',0))}ml | "
            f"{c.get('type','?')} | {c.get('gender','?')} | {c.get('price',0):.0f}ر.س"
            for j, c in enumerate(it["candidates"])
        )
        lines.append(f"[{i+1}] منتجنا: «{it['our']}» ({it['price']:.0f}ر.س)\n{cands}")

    prompt = (
        "خبير عطور فاخرة. لكل منتج اختر رقم المرشح المطابق تماماً أو 0 إذا لا يوجد.\n"
        "الشروط: نفس الماركة + نفس الحجم ±5ml + نفس EDP/EDT + نفس الجنس\n\n"
        + "\n\n".join(lines)
        + f'\n\nJSON فقط: {{"results":[r1,r2,...,r{len(batch)}]}}'
    )

    def _parse(txt):
        """يحلل استجابة AI إلى قائمة أرقام"""
        try:
            clean = re.sub(r'```json|```', '', txt).strip()
            s = clean.find('{'); e = clean.rfind('}') + 1
            if s < 0 or e <= s:
                return None
            raw = json.loads(clean[s:e]).get("results", [])
            out = []
            for j, it in enumerate(batch):
                # ✅ إصلاح #3: JSON مقتطع → 0 يُترجم لـ -1 (لا تطابق)
                n = raw[j] if j < len(raw) else 0
                try:
                    n = int(float(str(n)))
                except Exception:
                    n = 0
                if 1 <= n <= len(it["candidates"]):
                    out.append(n - 1)
                elif n == 0:
                    out.append(-1)
                else:
                    # ✅ إصلاح #3: رقم خارج النطاق → -1 لا index 0
                    out.append(-1)
            return out if len(out) == len(batch) else None
        except Exception:
            return None

    # ── 1. Gemini ─────────────────────────────────────────────────────────
    g_payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0, "maxOutputTokens": 500, "topP": 1, "topK": 1}
    }
    for key in (GEMINI_API_KEYS or []):
        if not key:
            continue
        if not _key_available(key):
            continue  # ⚡ مفتاح محظور — تخطّ فوراً
        try:
            r = _req.post(f"{_GURL}?key={key}", json=g_payload, timeout=12)
            if r.status_code == 200:
                txt = r.json()["candidates"][0]["content"]["parts"][0]["text"]
                out = _parse(txt)
                if out:
                    _cset(ck, out)
                    return out
            elif r.status_code == 429:
                _key_block(key)  # ⚡ احظر المفتاح 60 ثانية
                continue
            # 403/400 → جرب المفتاح التالي فوراً
        except Exception:
            continue

    # ── 2. OpenRouter fallback ────────────────────────────────────────────
    or_key = OPENROUTER_API_KEY
    if or_key:
        for model in _OR_FREE:
            try:
                r = _req.post(_OR_URL, json={
                    "model": model,
                    "messages": [{"role": "user", "content": prompt}],
                    "temperature": 0,
                    "max_tokens": 300,
                }, headers={
                    "Authorization": f"Bearer {or_key}",
                    "HTTP-Referer": "https://mahwous.com",
                }, timeout=30)
                if r.status_code == 200:
                    txt = r.json()["choices"][0]["message"]["content"]
                    out = _parse(txt)
                    if out:
                        _cset(ck, out)
                        return out
                elif r.status_code in (404, 400):
                    continue
                elif r.status_code in (401, 402):
                    break
            except Exception:
                continue

    # ── 3. Fuzzy fallback — لا يتوقف أبداً ──────────────────────────────
    # عند فشل كل AI → قرر حسب score الـ fuzzy
    # v31.7: خفض حد القبول من 88 إلى 82 لتسريع المعالجة
    out = []
    for it in batch:
        cands = it.get("candidates", [])
        if not cands:
            out.append(-1)
        elif cands[0].get("score", 0) >= 82:
            out.append(0)   # ثقة عالية → خذ الأول
        else:
            out.append(-1)  # ثقة منخفضة → مراجعة
    return out


# ═══════════════════════════════════════════════════════
#  صف «مستبعد» — لا يُفقد المنتج صمتاً عند فشل المطابقة (Zero Data Drop)
# ═══════════════════════════════════════════════════════
def _excluded_match_row(
    product,
    our_price,
    our_id,
    brand,
    size,
    ptype,
    gender,
    our_img="",
    our_url="",
    no="",
    *,
    score=0.0,
    مصدر_المطابقة="filtered_out",
):
    """يُرجع صفاً يظهر في التحليل بدل حذف المنتج عند عدم تجاوز فلاتر المطابقة."""
    sz_str = f"{int(size)}ml" if size else ""
    if score > 0:
        ai_lbl = f"⚪ مستبعد ({score:.0f}%)"
    else:
        ai_lbl = "⚪ مستبعد (لا يوجد تطابق)"
    return dict(
        المنتج=product or "غير معروف",
        معرف_المنتج=str(our_id or ""),
        السعر=float(our_price) if our_price is not None else 0.0,
        الماركة=str(brand or ""),
        الحجم=sz_str,
        النوع=ptype,
        الجنس=gender,
        NO=str(no or ""),
        منتج_المنافس="❌ لم يتجاوز فلاتر المطابقة / لا يوجد",
        معرف_المنافس="",
        سعر_المنافس=0.0,
        الفرق=0.0,
        نسبة_التطابق=float(score) if score else 0.0,
        ثقة_AI=ai_lbl,
        القرار="⚪ مستبعد (لا يوجد تطابق)",
        الخطورة="",
        المنافس="",
        عدد_المنافسين=0,
        جميع_المنافسين=[],
        مصدر_المطابقة=مصدر_المطابقة,
        تاريخ_المطابقة=datetime.now().strftime("%Y-%m-%d"),
        صورة_منتجنا=our_img or "",
        رابط_منتجنا=our_url or "",
        رابط_المنافس="",
    )


# ═══════════════════════════════════════════════════════
#  بناء صف النتيجة
# ═══════════════════════════════════════════════════════
def _row(product, our_price, our_id, brand, size, ptype, gender,
         best=None, override=None, src="", all_cands=None,
         our_img="", our_url="", no=""):
    sz_str = f"{int(size)}ml" if size else ""
    if best is None:
        return dict(المنتج=product, معرف_المنتج=our_id, السعر=our_price,
                    الماركة=brand, الحجم=sz_str, النوع=ptype, الجنس=gender,
                    NO=str(no or ""),
                    منتج_المنافس="—", معرف_المنافس="", سعر_المنافس=0,
                    الفرق=0, نسبة_التطابق=0, ثقة_AI="—",
                    القرار=override or "🔍 منتجات مفقودة",
                    الخطورة="", المنافس="", عدد_المنافسين=0,
                    جميع_المنافسين=[], مصدر_المطابقة=src or "—",
                    تاريخ_المطابقة=datetime.now().strftime("%Y-%m-%d"),
                    صورة_منتجنا=our_img or "", رابط_منتجنا=our_url or "",
                    رابط_المنافس="")

    cp    = float(best.get("price") or 0)
    score = float(best.get("score") or 0)
    diff  = round(our_price - cp, 2) if (our_price>0 and cp>0) else 0
    # نظام الخطورة حسب AI_COMPARISON_INSTRUCTIONS (نسبة مئوية + ثقة)
    diff_pct = abs((diff / cp) * 100) if cp > 0 else 0
    if diff_pct > 20 and score >= 85:
        risk = "🔴 حرج"
    elif diff_pct > 10 and score >= 75:
        risk = "🟡 متوسط"
    else:
        risk = "🟢 منخفض"

    # ═══ توزيع النتائج على الأقسام ═════════════════════════════════════
    # الحدود المستخدمة:
    #   score ≥ 85%           → مطابقة مؤكدة → توزيع سعري
    #   60% ≤ score < 85%     → تحت المراجعة (مطابقة محتملة)
    #   score < 60%           → صف «مستبعد» عبر _excluded_match_row (لا إخفاء صامت)
    # v31.6: استخدام قيم config.py
    NO_MATCH_THRESHOLD   = 60
    REVIEW_MAX           = MATCH_THRESHOLD if MATCH_THRESHOLD else 85

    # v31.6: حد سعري ديناميكي لتسامح «✅ موافق» حسب متوسط السعر (لا رقم سحري واحد لكل الفئات)
    def _smart_price_threshold(p1, p2):
        _DEFAULT_TOL = PRICE_TOLERANCE if PRICE_TOLERANCE else 10  # ر.س — تسامح افتراضي
        _HIGH_AVG, _HIGH_PCT = 300, 0.05   # سعر عالٍ (≥300): تسامح = 5% من المتوسط
        _MID_AVG = 100                      # سعر متوسط (≥100): تسامح ثابت = الافتراضي
        _LOW_TOL = 5                        # سعر منخفض (<100): تسامح صغير ثابت (ر.س)
        if p1 <= 0 or p2 <= 0:
            return _DEFAULT_TOL
        avg = (p1 + p2) / 2
        if avg >= _HIGH_AVG:   return avg * _HIGH_PCT
        elif avg >= _MID_AVG:  return _DEFAULT_TOL
        else:                  return _LOW_TOL

    if override:
        dec = override
    elif score < NO_MATCH_THRESHOLD:
        return _excluded_match_row(
            product, our_price, our_id, brand, size, ptype, gender,
            our_img=our_img, our_url=our_url, no=no,
            score=score,
            مصدر_المطابقة="score_below_60",
        )
    elif src in ("gemini","auto") or score >= REVIEW_MAX:
        # مطابقة مؤكدة (≥85%) → توزيع حسب السعر
        if our_price > 0 and cp > 0:
            _pt = _smart_price_threshold(our_price, cp)
            if diff > _pt:       dec = "🔴 سعر أعلى"
            elif diff < -_pt:    dec = "🟢 سعر أقل"
            else:                dec = "✅ موافق"
        else:
            dec = "⚠️ تحت المراجعة — بلا سعر منافس"  # P5: مطابَق مؤكّد بلا سعر → مراجعة لا «مفقود»
    else:
        # 60% ≤ score < 85% → مطابقة محتملة → تحت المراجعة
        dec = f"⚠️ تحت المراجعة ({score:.0f}%)"

    ai_lbl = {"gemini":f"🤖✅({score:.0f}%)",
              "auto":f"🎯({score:.0f}%)",
              "gemini_no_match":"🤖❌"}.get(src, f"{score:.0f}%")

    # ── بناء قائمة المنافسين المتطابقين فعلاً (score ≥ NO_MATCH_THRESHOLD) ──────
    # لا تُدرج في البطاقة أي منافس لم تثبت مطابقته — منع عرض منتجات غير متطابقة
    confirmed_cands = [
        c for c in (all_cands or [best])
        if isinstance(c, dict) and float(c.get("score", 0) or 0) >= NO_MATCH_THRESHOLD
    ]
    if not confirmed_cands:
        confirmed_cands = [best]

    # إزالة التكرار — منافس واحد فقط في القائمة (أفضل/أرخص منتج لديه)
    unique_competitors = []
    seen_competitors: set = set()
    seen_candidates: set = set()
    for cand in confirmed_cands:
        comp_name = str(cand.get("competitor", "") or "").strip()
        prod_name = str(cand.get("name", "") or "").strip()
        prod_id   = str(cand.get("product_id", "") or "").strip()
        prod_url  = str(cand.get("product_url") or cand.get("url") or "").strip()
        candidate_key = (
            comp_name.lower(),
            prod_id or prod_url or normalize(prod_name),
        )
        if candidate_key in seen_candidates:
            continue
        seen_candidates.add(candidate_key)
        comp_key = comp_name.lower()
        if comp_key and comp_key not in seen_competitors:
            unique_competitors.append(cand)
            seen_competitors.add(comp_key)
        elif not comp_name and not unique_competitors:
            unique_competitors.append(cand)

    ac = unique_competitors[:10] or [best]

    # ── اختيار بطل البطاقة الرئيسية: الأرخص سعراً بين المتطابقين ──────────────
    # البطاقة الرئيسية (VS card) تعرض: منتجنا VS أقل منافس سعراً
    # هذا هو المنافس الذي يؤثر فعلاً على قرار التسعير
    cheapest = min(
        (c for c in ac if float(c.get("price", 0) or 0) > 0),
        key=lambda c: float(c.get("price", 0) or 0),
        default=best,
    )

    # أعد حساب السعر والفرق بناءً على الأرخص
    cp_display   = float(cheapest.get("price") or 0)
    diff_display = round(our_price - cp_display, 2) if (our_price > 0 and cp_display > 0) else 0
    score_display = float(cheapest.get("score") or score)

    # أعد حساب الخطورة بناءً على الأرخص (المنافس الذي يضغط فعلاً على السعر)
    diff_pct_display = abs((diff_display / cp_display) * 100) if cp_display > 0 else 0
    if diff_pct_display > 20 and score_display >= 85:
        risk = "🔴 حرج"
    elif diff_pct_display > 10 and score_display >= 75:
        risk = "🟡 متوسط"
    else:
        risk = "🟢 منخفض"

    # أعد حساب القرار بناءً على الأرخص
    if not override:
        # score < NO_MATCH_THRESHOLD مُعالَج ومُرجَع أعلاه (السطر ~2333) → لا يصل هنا
        if src in ("gemini", "auto") or score >= REVIEW_MAX:
            # أ3 — بوابة التحقق الشديد: لا بطاقة إلا بعد فحص هيكلي (حجم/ماركة).
            # يعيد استخدام نفس منطق المفقودات المُختبَر (_structural_match): يرفض فقط
            # عند اختلاف حجم/ماركة مؤكّد، وإلّا يقبل (لا عقاب عند الغموض → صفر فقدان).
            # استيراد محلي لتفادي أي استيراد دائري عند تحميل الوحدة.
            from engines.missing_products_engine import _structural_match as _struct_ok
            if not _struct_ok(str(cheapest.get("name", "") or ""), product):
                dec = f"⚠️ تحت المراجعة — اختلاف هيكلي ({score_display:.0f}%)"
            elif our_price > 0 and cp_display > 0:
                _pt2 = _smart_price_threshold(our_price, cp_display)
                if diff_display > _pt2:   dec = "🔴 سعر أعلى"
                elif diff_display < -_pt2: dec = "🟢 سعر أقل"
                else:                      dec = "✅ موافق"

    # ترتيب comp_strip من الأرخص للأغلى (مرئياً)
    ac_sorted = sorted(ac, key=lambda c: float(c.get("price", 0) or 0))

    competitor_names = sorted({
        str(c.get("competitor", "") or "").strip()
        for c in ac_sorted
        if str(c.get("competitor", "") or "").strip()
    })
    best_competitor = str(cheapest.get("competitor", "") or "").strip()
    if not best_competitor and competitor_names:
        best_competitor = competitor_names[0]

    return dict(المنتج=product, معرف_المنتج=our_id, السعر=our_price,
                الماركة=brand, الحجم=sz_str, النوع=ptype, الجنس=gender,
                NO=str(no or ""),
                منتج_المنافس=cheapest["name"], معرف_المنافس=cheapest.get("product_id", ""),
                سعر_المنافس=cp_display, الفرق=diff_display,
                نسبة_التطابق=score_display, ثقة_AI=ai_lbl,
                القرار=dec, الخطورة=risk, المنافس=best_competitor,
                عدد_المنافسين=len(competitor_names),
                جميع_المنافسين=ac_sorted, مصدر_المطابقة=src or "fuzzy",
                تاريخ_المطابقة=datetime.now().strftime("%Y-%m-%d"),
                صورة_منتجنا=our_img or "", رابط_منتجنا=our_url or "",
                رابط_المنافس=str(cheapest.get("product_url") or cheapest.get("url") or "").strip())


# ═══════════════════════════════════════════════════════
#  مطابقة منتج واحد — مساعد للـ Real-Time Pipeline (Task 2.3)
# ═══════════════════════════════════════════════════════

def match_single_product(
    product: str,
    our_price: float,
    our_id: str,
    brand: str,
    display_brand: str,
    size: float,
    ptype: str,
    gender: str,
    our_img: str,
    our_url: str,
    indices: dict,
    use_ai: bool = False,
) -> dict:
    """
    Match one of our products against pre-built CompIndex objects.

    Extracted from run_full_analysis() so the real-time pipeline can call it
    per-row without re-building the indices on every call.

    Args:
        product      : normalised product name (str)
        our_price    : our selling price (float)
        our_id       : our product id / SKU (str)
        brand        : brand extracted from product name
        display_brand: brand from row data (preferred) or extracted brand
        size         : size in ml (float, 0 if unknown)
        ptype        : product type string (EDP / EDT / …)
        gender       : gender label (رجالي / نسائي / للجنسين)
        our_img      : our product image URL
        our_url      : our product page URL
        indices      : dict {competitor_name: CompIndex} — pre-built
        use_ai       : if True and score in 60-96 range, calls _ai_batch inline

    Returns:
        A result dict with the same schema as _row() / _excluded_match_row().
        Never returns None — always returns a complete result row.
    """
    our_n  = normalize(product)
    our_pl = extract_product_line(product, brand)

    # Gather candidates from all competitor indices
    all_cands: list = []
    for idx_obj in indices.values():
        all_cands.extend(
            idx_obj.search(our_n, brand, size, ptype, gender,
                           our_pline=our_pl, top_n=6, our_price=our_price)
        )

    if not all_cands:
        return _excluded_match_row(
            product, our_price, our_id, display_brand, size, ptype, gender,
            our_img=our_img, our_url=our_url,
            score=0.0,
            مصدر_المطابقة="no_candidates",
        )

    all_cands.sort(key=lambda x: x["score"], reverse=True)
    top5  = all_cands[:5]
    best0 = top5[0]

    if best0["score"] < 60:
        return _excluded_match_row(
            product, our_price, our_id, display_brand, size, ptype, gender,
            our_img=our_img, our_url=our_url,
            score=float(best0.get("score") or 0),
            مصدر_المطابقة="below_match_threshold",
        )

    # High-confidence auto match — no AI needed
    if best0["score"] >= 97 or not use_ai:
        result = _row(
            product, our_price, our_id, brand, size, ptype, gender,
            best0, src="auto", all_cands=all_cands,
            our_img=our_img, our_url=our_url,
        )
        return result if result is not None else _excluded_match_row(
            product, our_price, our_id, display_brand, size, ptype, gender,
            our_img=our_img, our_url=our_url,
            score=float(best0.get("score") or 0),
            مصدر_المطابقة="auto_none",
        )

    # Mid-confidence — call AI inline (single-item batch)
    pending_item = dict(
        product=product, our_price=our_price, our_id=our_id,
        brand=brand, size=size, ptype=ptype, gender=gender,
        candidates=top5, all_cands=all_cands,
        our=product, price=our_price,
        our_img=our_img, our_url=our_url,
    )
    try:
        idxs = _ai_batch([pending_item])
        ci   = idxs[0] if idxs else -1
    except Exception:
        ci = -1  # AI failed — fall back to best fuzzy

    if ci < 0:
        best_fb = top5[0]
        result  = _row(
            product, our_price, our_id, brand, size, ptype, gender,
            best_fb, src="ai_uncertain", all_cands=all_cands,
            our_img=our_img, our_url=our_url,
        )
    else:
        result = _row(
            product, our_price, our_id, brand, size, ptype, gender,
            top5[ci], src="gemini", all_cands=all_cands,
            our_img=our_img, our_url=our_url,
        )

    return result if result is not None else _excluded_match_row(
        product, our_price, our_id, display_brand, size, ptype, gender,
        our_img=our_img, our_url=our_url,
        score=float(best0.get("score") or 0),
        مصدر_المطابقة="ai_result_none",
    )


# ═══════════════════════════════════════════════════════
#  التحليل الكامل — v21 الهجين الفائق السرعة
# ═══════════════════════════════════════════════════════
def run_full_analysis(our_df, comp_dfs, progress_callback=None, use_ai=True,
                      ledger=None):
    """
    1. بناء CompIndex لكل منافس (تطبيع مسبق)
    2. لكل منتجنا → search vectorized
    3. score≥97 → تلقائي | 62-96 → AI batch | <62 → مراجعة

    يُرجع: (DataFrame النتائج, audit_stats)

    Phase 0: if ``ledger`` is provided (observability.CompetitorIntakeLedger),
    every competitor row is recorded at ingest, every match-decision site
    writes a terminal state, and ``audit_stats["ledger"]`` reports the
    end-of-run counters + invariant check.
    """
    import traceback as _tb
    from observability.ledger import (
        NullLedger, state_from_status, ingest_comp_df,
        make_comp_id, CONFIRMED_MATCH, REJECTED_STRUCTURAL,
        REJECTED_LOW_CONFIDENCE, ERROR,
    )
    _led = ledger if ledger is not None else NullLedger()

    results = []
    audit_stats = {
        "total_input": int(len(our_df)) if our_df is not None else 0,
        "processed": 0,
        "skipped_empty": 0,
        "skipped_samples": 0,
        "no_competitor_found": 0,
        "excluded_sets": 0,
    }
    our_col       = _name_col_for_analysis(our_df)
    our_price_col = _fcol(our_df, ["سعر المنتج","السعر","سعر","Price","price","PRICE"])
    our_id_col    = _fcol_optional(our_df, [
        "رقم المنتج","معرف المنتج","المعرف","معرف","رقم_المنتج","معرف_المنتج",
        "product_id","Product ID","Product_ID","ID","id","Id",
        "SKU","sku","Sku","رمز المنتج","رمز_المنتج","رمز المنتج sku",
        "الكود","كود","Code","code","الرقم","رقم","Barcode","barcode","الباركود"
    ]) or ""
    our_img_col = _fcol_optional(our_df, [
        "صورة المنتج", "صوره المنتج", "image", "Image", "product_image", "الصورة",
    ])
    our_url_col = _fcol_optional(our_df, [
        "رابط المنتج", "الرابط", "رابط", "product_url", "link", "url", "URL",
    ])
    our_brand_col = _fcol_optional(our_df, ["الماركة", "Brand", "brand", "البراند"])
    # رقم No. من كتالوجنا — Primary Key المُرسَل لـ Make/سلة عند تحديث السعر
    our_no_col = _fcol_optional(our_df, ["No.", "NO", "no", "No"])

    # ⚡ v22: Enrich brand vocabulary with competitor brands BEFORE building indices
    enrich_known_brands(comp_dfs=comp_dfs)

    # ── بناء الفهارس المسبقة ──
    indices = {}
    # Phase 0: remember (competitor, url) → comp_id so transition sites can
    # reconstruct the id from the cand dict without piggy-backing on CompIndex.
    for cname, cdf in comp_dfs.items():
        ccol = _name_col_for_analysis(cdf)
        icol = _fcol_optional(cdf, [
            "رقم المنتج","معرف المنتج","المعرف","معرف","رقم_المنتج","معرف_المنتج",
            "product_id","Product ID","Product_ID","ID","id","Id",
            "SKU","sku","Sku","رمز المنتج","رمز_المنتج","رمز المنتج sku",
            "الكود","كود","Code","code","الرقم","رقم","Barcode","barcode","الباركود"
        ]) or ""
        c_img = _fcol_optional(cdf, [
            "صورة المنتج", "صوره المنتج", "image_url", "image", "Image", "product_image", "الصورة",
        ])
        c_url = _fcol_optional(cdf, [
            "رابط المنتج", "الرابط", "رابط", "product_url", "link", "url", "URL",
        ])
        indices[cname] = CompIndex(cdf, ccol, icol, cname, img_col=c_img, url_col=c_url)
        # Phase 0: ingest every competitor row before any filter runs.
        try:
            ingest_comp_df(_led, cname, cdf, ccol, url_col=c_url or "")
        except Exception as _ie:
            # Never let instrumentation break the run; just log and continue.
            import logging as _lg
            _lg.getLogger("engines.engine").warning(
                "ledger ingest error for %s: %s", cname, _ie,
            )

    def _cand_comp_id(cand):
        """Rebuild the ledger comp_id for a candidate returned by CompIndex."""
        if not cand:
            return None
        return make_comp_id(
            cand.get("competitor", ""),
            cand.get("name", ""),
            cand.get("product_url", "") or "",
        )

    total   = len(our_df)
    pending = []
    BATCH   = 30  # v31.7: حجم دفعة أكبر = API calls أقل = أسرع بكثير

    # v31.7: معالجة AI متوازية — 3 خيوط متزامنة
    from concurrent.futures import ThreadPoolExecutor, as_completed

    def _flush():
        """يُعالج الـ pending batch ويضيف النتائج مباشرة — محمي من الأخطاء"""
        if not pending:
            return
        try:
            idxs = _ai_batch(pending)
        except Exception:
            # فشل AI → fallback: استخدم أفضل مرشح fuzzy
            idxs = []
            for it in pending:
                cands = it.get("candidates", [])
                if cands and cands[0].get("score", 0) >= 88:
                    idxs.append(0)
                else:
                    idxs.append(-1)
        for j, it in enumerate(pending):
            try:
                # ✅ إصلاح: idxs أقصر → -1 (لا تطابق) لا index 0
                ci = idxs[j] if j < len(idxs) else -1
                if ci < 0:
                    # AI غير متأكد → أعطِ أفضل مرشح كمراجعة
                    best_fallback = it["candidates"][0] if it["candidates"] else None
                    rr = _row(it["product"], it["our_price"], it["our_id"],
                              it["brand"], it["size"], it["ptype"], it["gender"],
                              best_fallback, "⚠️ تحت المراجعة", "ai_uncertain",
                              all_cands=it["all_cands"],
                              our_img=it.get("our_img", ""), our_url=it.get("our_url", ""),
                              no=it.get("our_no", ""))
                    _cid = _cand_comp_id(best_fallback)
                    if _cid:
                        _led.mark_state(_cid, CONFIRMED_MATCH,
                                        reason_code="under_review",
                                        last_score=float(best_fallback.get("score") or 0) if best_fallback else None)
                else:
                    best = it["candidates"][ci]
                    rr = _row(it["product"], it["our_price"], it["our_id"],
                              it["brand"], it["size"], it["ptype"], it["gender"],
                              best, src="gemini", all_cands=it["all_cands"],
                              our_img=it.get("our_img", ""), our_url=it.get("our_url", ""),
                              no=it.get("our_no", ""))
                    _cid = _cand_comp_id(best)
                    if _cid:
                        _led.mark_state(_cid, CONFIRMED_MATCH,
                                        reason_code="ai_match",
                                        last_score=float(best.get("score") or 0))
                if rr is not None:
                    results.append(rr)
            except Exception as _flush_exc:
                # Phase 0: never drop silently. Record an error row in the
                # ledger so the invariant still balances and the run report
                # shows where the loss would have happened.
                _best = it["candidates"][0] if it.get("candidates") else None
                _cid = _cand_comp_id(_best)
                _led.mark_error(
                    _cid, "flush_row_error",
                    _tb.format_exc()[:300] if hasattr(_tb, "format_exc") else str(_flush_exc),
                )
                import logging as _lg
                _lg.getLogger("engines.engine").error(
                    "flush row error (comp_id=%s): %s", _cid, _flush_exc,
                )
                continue
        pending.clear()
        # ✅ إصلاح #2: حذف time.sleep(0.5) — مخالف لقواعد Streamlit Main Thread

    def _cell_clean(r, col):
        if not col or col not in r.index:
            return ""
        v = r.get(col, "")
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ""
        s = str(v).strip()
        if s.lower() in ("nan", "none", "<na>"):
            return ""
        return s

    for i, (_, row) in enumerate(our_df.iterrows()):
        product = str(row.get(our_col, "")).strip()
        our_no  = _cell_clean(row, our_no_col)
        if not product:
            audit_stats["skipped_empty"] += 1
            # v31.6: لا نفقد المنتج — نضيفه في النتائج
            results.append(_excluded_match_row(
                "(اسم فارغ)", 0.0, _pid(row, our_id_col), "", 0, "", "",
                our_img=_cell_clean(row, our_img_col),
                our_url=_cell_clean(row, our_url_col),
                no=our_no,
                score=0, مصدر_المطابقة="skipped_empty_name",
            ))
            if progress_callback:
                progress_callback((i + 1) / total, results)
            continue

        # v31.6: التستر والعينات لم تعد تُسقط — تُقارن مع نظيراتها
        _product_class = classify_product(product)

        # v31.11b: المجموعات والطقم لا تُقارن مع عطور فردية — تُستبعد مباشرة
        _is_our_set = any(w in product.lower() for w in ('مجموعة','مجموعه','طقم'))
        if _is_our_set or _product_class == 'set':
            audit_stats["excluded_sets"] += 1
            results.append(_excluded_match_row(
                product, 0.0, _pid(row, our_id_col), "", extract_size(product), "", "",
                our_img=_cell_clean(row, our_img_col),
                our_url=_cell_clean(row, our_url_col),
                no=our_no,
                score=0, مصدر_المطابقة="excluded_set_product",
            ))
            if progress_callback:
                progress_callback((i + 1) / total, results)
            continue

        size_ml = extract_size(product)
        if size_ml > 0 and size_ml < 2:
            # أقل من 2ml فقط — عينات صغيرة جداً
            audit_stats["skipped_samples"] += 1
            results.append(_excluded_match_row(
                product, 0.0, _pid(row, our_id_col), "", size_ml, "", "",
                our_img=_cell_clean(row, our_img_col),
                our_url=_cell_clean(row, our_url_col),
                no=our_no,
                score=0, مصدر_المطابقة="skipped_micro_size",
            ))
            if progress_callback:
                progress_callback((i + 1) / total, results)
            continue

        our_price = 0.0
        if our_price_col:
            try:
                our_price = float(str(row[our_price_col]).replace(",", ""))
            except Exception:
                pass

        our_id  = _pid(row, our_id_col)
        our_img = _cell_clean(row, our_img_col)
        our_url = _cell_clean(row, our_url_col)
        brand   = extract_brand(product)
        brand_from_row = _cell_clean(row, our_brand_col) if our_brand_col else ""
        display_brand = brand_from_row or brand
        size    = size_ml
        ptype   = extract_type(product)
        gender  = extract_gender(product)
        our_n   = normalize(product)
        our_pl  = extract_product_line(product, brand)

        # ── جمع المرشحين من كل الفهارس ──
        all_cands = []
        for idx_obj in indices.values():
            all_cands.extend(idx_obj.search(our_n, brand, size, ptype, gender,
                                            our_pline=our_pl, top_n=3, our_price=our_price))

        # v31.11b: تنقية all_cands — حذف المنتجات بخط إنتاج مختلف
        if our_pl and all_cands:
            _cleaned = []
            for c in all_cands:
                c_name = c.get("name", "")
                c_br   = c.get("brand", "")
                if c_br and brand and normalize(c_br).lower() == normalize(brand).lower():
                    c_pl = extract_product_line(c_name, c_br)
                    if c_pl and our_pl:
                        _sim = fuzz.token_sort_ratio(our_pl.lower(), c_pl.lower())
                        if _sim < 55:
                            continue  # خط منتج مختلف → حذف من البطاقة
                _cleaned.append(c)
            all_cands = _cleaned if _cleaned else all_cands[:1]

        if not all_cands:
            audit_stats["no_competitor_found"] += 1
            results.append(
                _excluded_match_row(
                    product, our_price, our_id, display_brand, size, ptype, gender,
                    our_img=our_img, our_url=our_url, no=our_no,
                    score=0.0,
                    مصدر_المطابقة="no_candidates",
                )
            )
            # Phase 0: no competitor matched our product at all — there is no
            # comp row to mark here; the sweep will handle untouched comp rows
            # at end-of-run. Nothing to do on the ledger side for this branch.
            if progress_callback:
                progress_callback((i + 1) / total, results)
            continue

        audit_stats["processed"] += 1

        all_cands.sort(key=lambda x: x["score"], reverse=True)
        top5  = all_cands[:5]
        best0 = top5[0]

        if best0["score"] < 60:
            results.append(
                _excluded_match_row(
                    product, our_price, our_id, display_brand, size, ptype, gender,
                    our_img=our_img, our_url=our_url, no=our_no,
                    score=float(best0.get("score") or 0),
                    مصدر_المطابقة="below_match_threshold",
                )
            )
            _cid = _cand_comp_id(best0)
            if _cid:
                _led.mark_state(_cid, REJECTED_LOW_CONFIDENCE,
                                reason_code="below_match_threshold",
                                last_score=float(best0.get("score") or 0))
            if progress_callback:
                progress_callback((i + 1) / total, results)
            continue

        if best0["score"] >= 85:  # ≥85% → مطابقة تلقائية مؤكدة (بطاقة)
            row_result = _row(product, our_price, our_id, brand, size, ptype, gender,
                              best0, src="auto", all_cands=all_cands,
                              our_img=our_img, our_url=our_url, no=our_no)
            if row_result is not None:   # ← فلتر None
                results.append(row_result)
                _cid = _cand_comp_id(best0)
                if _cid:
                    _led.mark_state(_cid, CONFIRMED_MATCH,
                                    reason_code="auto_match",
                                    last_score=float(best0.get("score") or 0))
            else:
                # v31.6: safety net
                audit_stats["dropped_none"] = audit_stats.get("dropped_none", 0) + 1
                results.append(_excluded_match_row(
                    product, our_price, our_id, brand, size, ptype, gender,
                    our_img=our_img, our_url=our_url, no=our_no,
                    score=float(best0.get("score") or 0),
                    مصدر_المطابقة="row_returned_none",
                ))
        elif use_ai:
            pending.append(dict(
                product=product, our_price=our_price, our_id=our_id,
                brand=brand, size=size, ptype=ptype, gender=gender,
                candidates=top5, all_cands=all_cands,
                our=product, price=our_price,
                our_img=our_img, our_url=our_url, our_no=our_no,
            ))
            if len(pending) >= BATCH:
                _flush()
        else:
            # إصلاح المهلة: بلا AI، النطاق الرمادي (60-84%) → مراجعة يدوية
            # (لا بطاقة على ثقة منخفضة، لا فقدان بيانات). AI يُشغَّل يدوياً على قسم المراجعة.
            row_result = _row(product, our_price, our_id, brand, size, ptype, gender,
                              best0, "⚠️ تحت المراجعة", "no_ai_review",
                              all_cands=all_cands,
                              our_img=our_img, our_url=our_url, no=our_no)
            if row_result is not None:
                results.append(row_result)
                _cid = _cand_comp_id(best0)
                if _cid:
                    _led.mark_state(_cid, CONFIRMED_MATCH,
                                    reason_code="under_review",
                                    last_score=float(best0.get("score") or 0))

        if progress_callback:
            progress_callback((i + 1) / total, results)

    _flush()

    # ── Phase 0: end-of-run ledger sweep + invariant check ───────────────
    try:
        swept = _led.sweep_untransitioned(
            default_state=REJECTED_LOW_CONFIDENCE,
            reason_code="not_selected_in_batch",
        )
        ok, report = _led.check_invariant()
        audit_stats["ledger"] = report
        audit_stats["ledger_sweep_count"] = swept
        if not ok:
            import logging as _lg
            _lg.getLogger("engines.engine").error(
                "pipeline invariant FAILED: %s", report,
            )
    except Exception as _le:
        import logging as _lg
        _lg.getLogger("engines.engine").warning(
            "ledger finalize error (non-fatal): %s", _le,
        )

    # ── تنظيف الذاكرة بعد المعالجة الثقيلة ──────────────────────────────
    _out = pd.DataFrame(results)
    del results
    del indices
    del pending
    gc.collect()

    return _out, audit_stats


# ═══════════════════════════════════════════════════════
#  تصنيف المنتج: عطور / عناية / تجميل / أخرى
# ═══════════════════════════════════════════════════════
def classify_product_category(name: str) -> str:
    """تصنيف المنتج حسب اسمه إلى: عطور / عناية / تجميل / أخرى"""
    if not name:
        return "📦 أخرى"
    n = name.lower()
    # عطور
    if any(kw in n for kw in (
        'عطر', 'بارفيوم', 'بيرفيوم', 'كولون', 'او دو', 'edp', 'edt',
        'perfume', 'cologne', 'fragrance', 'parfum', 'مسك', 'بخور', 'عود',
        'دهن', 'spray', 'بودي ميست', 'body mist', 'تستر', 'tester',
        'eau de', 'او دي', 'اكستريت', 'extrait', 'انتنس', 'intense',
    )):
        return "🌸 عطور"
    # تجميل
    if any(kw in n for kw in (
        'أحمر شفاه', 'احمر شفاه', 'كريم أساس', 'كريم اساس', 'ماسكارا',
        'آيلاينر', 'ايلاينر', 'بودرة', 'مكياج', 'كونسيلر', 'هايلايتر',
        'بلاشر', 'ظلال', 'روج', 'ملمع شفاه', 'برايمر',
        'foundation', 'lipstick', 'mascara', 'concealer', 'makeup',
        'eyeliner', 'blush', 'highlighter', 'primer', 'lip gloss',
    )):
        return "💄 تجميل"
    # عناية
    if any(kw in n for kw in (
        'شامبو', 'بلسم', 'مرطب', 'واقي شمس', 'غسول', 'سيروم',
        'لوشن', 'تونر', 'قناع', 'ماسك', 'كريم مرطب', 'مقشر',
        'serum', 'moisturizer', 'shampoo', 'sunscreen', 'lotion',
        'cleanser', 'scrub', 'بشرة', 'للشعر', 'للجسم', 'صابون',
        'زيت شعر', 'زيت جسم', 'ديودرانت', 'deodorant', 'shower gel',
        'body lotion', 'hand cream', 'كريم يد',
    )):
        return "🧴 عناية"
    return "📦 أخرى"


# ═══════════════════════════════════════════════════════
#  المنتجات المفقودة — كشف التكرار الفائق الدقة v22
# ═══════════════════════════════════════════════════════
def find_missing_products(our_df, comp_dfs):
    """
    v26 — كشف المنتجات المفقودة الفائق الدقة:
    ✅ 5 خوارزميات تشابه + مطابقة بالكلمات
    ✅ كشف تستر↔أساسي (badge) — لا ضياع فرص
    ✅ تطبيع شامل للأسماء العربية والإنجليزية
    ✅ حد ثقة مزدوج: موجود(82%) / مشابه(68%)
    ✅ منع التكرار من منافسين مختلفين
    """
    our_col = _name_col_for_analysis(our_df)

    # ── بناء فهرس منتجاتنا الكامل ─────────────────────────────────────
    our_items = []
    for _, r in our_df.iterrows():
        name = str(r.get(our_col, "")).strip()
        if not name or is_sample(name): continue
        brand  = extract_brand(name)
        norm   = normalize(name)
        # normalize_aggressive: يحذف عطر/بارفيوم/بيرفيوم... للمطابقة الحساسة
        agg    = normalize_name(name)   # ← normalize_name
        pline  = extract_product_line(name, brand)
        is_t   = is_tester(name)
        # نسخة مُجرَّدة من "تستر" للمقارنة مع الأساسي
        bare_n    = re.sub(r"\btester\b|تستر|tester", "", agg).strip()
        our_items.append({
            "raw":      name,
            "norm":     norm,
            "agg":      agg,       # ← النسخة العنيفة للمطابقة
            "bare":     bare_n,    # ← بدون تستر
            "brand":    brand,
            "pline":    pline,
            "size":     extract_size(name),
            "type":     extract_type(name),
            "gender":   extract_gender(name),
            "is_tester": is_t,
        })

    # ── فهرس سريع بالكلمات (مبني على agg المطبَّع عنيفاً) ──────────────
    _word_idx = {}
    for p in our_items:
        for w in set(p["bare"].split()):
            if len(w) >= 3:  # ← 3 بدل 4 لاستيعاب كلمات عربية قصيرة
                _word_idx.setdefault(w, []).append(p)

    def _word_overlap(a, b):
        sa = set(a.split()); sb = set(b.split())
        if not sa or not sb: return 0
        return len(sa & sb) / len(sa | sb) * 100

    def _score_pair(cn, on, c_pline, o_pline):
        """
        cn/on هما النسختان العنيفتان (normalize_aggressive).
        3 خوارزميات مرجحة: token_set (الأقوى) + token_sort + partial.
        """
        s1 = fuzz.token_sort_ratio(cn, on)    # يتجاهل الترتيب
        s2 = fuzz.token_set_ratio(cn, on)     # الأقوى: يتجاهل الكلمات الزائدة
        s3 = fuzz.partial_ratio(cn, on)       # يجد نصاً ضمن نص
        base = s1*0.30 + s2*0.50 + s3*0.20   # token_set له وزن أعلى
        s5 = fuzz.token_set_ratio(c_pline, o_pline) if (c_pline and o_pline) else 0
        return base, s2, s5

    def _get_candidates(bare_cn):
        """فهرس الكلمات للبحث السريع — يستخدم bare (normalize_aggressive بدون تستر)"""
        seen = {}
        for w in set(bare_cn.split()):
            if len(w) >= 3 and w in _word_idx:
                for p in _word_idx[w]:
                    seen[id(p)] = p
        # fallback: إذا لم يجد شيئاً → ابحث في كامل القائمة
        return list(seen.values()) if seen else our_items

    def _is_same_product(cp_raw, cn, c_brand, c_pline, c_size, c_type, c_gender, c_is_tester, c_agg=""):
        """
        يُعيد: (found, score, reason, variant_info)
        variant_info = None | {"type":"tester"|"base","product":p,"score":float}
        cn   = normalize(cp_raw)   — للمعلومات المساعدة
        c_agg= normalize_aggressive(cp_raw) — للمقارنة الفعلية
        """
        if not c_agg:
            c_agg = normalize_name(cp_raw)  # ← normalize_name
        bare_cn = re.sub(r"\btester\b|تستر|tester", "", c_agg).strip()
        c_brand_n = normalize(c_brand) if c_brand else ""

        # فرز المرشحين: نفس الماركة أولاً
        candidates = _get_candidates(bare_cn)
        if c_brand_n:
            priority = [p for p in candidates if normalize(p["brand"]) == c_brand_n]
            others   = [p for p in candidates if normalize(p["brand"]) != c_brand_n]
            candidates = priority + others[:100]

        best_same   = (0, None, "")
        best_variant= (0, None, "")   # تستر ↔ أساسي

        for p in candidates[:400]:
            # ← المقارنة على bare (agg بدون تستر) بدل norm
            o_bare = p["bare"]
            base, set_sc, pline_sc = _score_pair(bare_cn, o_bare, c_pline, p["pline"])

            # ── عقوبات ──────────────────────────────────────────────
            penalty = 0
            if c_size > 0 and p["size"] > 0:
                d = abs(c_size - p["size"])
                if d > 50: penalty += 35
                elif d > 20: penalty += 22
                elif d > 8:  penalty += 12
            if c_type and p["type"] and c_type != p["type"]: penalty += 12
            if c_gender and p["gender"] and c_gender != p["gender"]: penalty += 40
            if c_pline and p["pline"]:
                pl = fuzz.token_sort_ratio(c_pline, p["pline"])
                if pl < 60: penalty += 30
                elif pl < 75: penalty += 18
                elif pl < 88: penalty += 8
            if c_brand_n and p["brand"] and normalize(p["brand"]) == c_brand_n:
                base += 5

            final = max(0, min(100, base - penalty))

            # هل نفس النوع (كلاهما تستر أو كلاهما أساسي)؟
            same_type = (p["is_tester"] == c_is_tester)

            if same_type:
                if final > best_same[0]:
                    best_same = (final, p, f"يشبه «{p['raw'][:50]}» ({final:.0f}%)")
                if final >= 95:
                    return True, final, best_same[2], None
            else:
                if final > best_variant[0]:
                    best_variant = (final, p, f"{'تستر' if p['is_tester'] else 'العطر الأساسي'}")

        # ── قرار النوع المطابق ─────────────────────────────────────────
        # بعد normalize_aggressive: 75% كافية للتأكد (الضجيج محذوف)
        CONFIRMED = 82   # ← رُفِع إلى 82%: فقط ثقة 82%+ أننا نملكه تُستبعد من المفقودة
        SIMILAR   = 70   # ← رُفِع إلى 70%: حد "مشابه محتمل" — يظهر للمستخدم مع تحذير

        if best_same[0] >= CONFIRMED:
            return True, best_same[0], best_same[2], None
        if best_same[0] >= SIMILAR:
            # منطقة رمادية → مفقود لكن مع تحذير للمستخدم
            vinfo = {"type": "similar",
                     "product": best_same[1]["raw"] if best_same[1] else "",
                     "score": best_same[0]} if best_same[1] else None
            return False, best_same[0], f"⚠️ مشابه ({best_same[0]:.0f}%) — {best_same[2]}", vinfo

        # ── كشف التستر/الأساسي ───────────────────────────────────────
        variant_info = None
        if best_variant[0] >= 55 and best_variant[1]:
            p_var  = best_variant[1]
            v_type = "tester" if p_var["is_tester"] else "base"
            variant_info = {
                "type":    v_type,
                "label":   "🏷️ يتوفر لدينا تستر منه" if v_type == "tester" else "✅ يتوفر لدينا العطر الأساسي",
                "product": p_var["raw"],
                "score":   best_variant[0],
            }

        return False, best_same[0], "", variant_info

    # ── البحث الرئيسي ─────────────────────────────────────────────────
    missing  = []
    seen_bare = set()   # مفاتيح إزالة التكرار داخل نفس المنافس فقط
    # عدّادات أسباب الاستبعاد (الإصلاح 1 + تقرير التشخيص)
    _drop_nosize = 0
    _drop_mini   = 0
    _MIN_SIZE_ML = 10.0

    def _cell_str(r, col):
        if not col or col not in r.index:
            return ""
        v = r.get(col, "")
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ""
        s = str(v).strip()
        if s.lower() in ("nan", "none", "<na>"):
            return ""
        return s

    for cname, cdf in comp_dfs.items():
        ccol = _name_col_for_analysis(cdf)
        # لا نستخدم _fcol للمعرّف — كان يسقط إلى العمود الأول فيُعرَض اسم المنتج كـ SKU
        icol = _fcol_optional(cdf, [
            "رقم المنتج","معرف المنتج","المعرف","معرف","رقم_المنتج","معرف_المنتج",
            "product_id","Product ID","Product_ID","ID","id","Id",
            "SKU","sku","Sku","رمز المنتج","رمز_المنتج",
            "الكود","كود","Code","code","الرقم","رقم","Barcode","barcode","الباركود"
        ]) or ""
        img_col = _find_image_column(cdf) or ""
        url_col = _find_url_column(cdf) or ""

        for _, row in cdf.iterrows():
            cp = str(row.get(ccol, "")).strip()
            if not cp or is_sample(cp): continue

            cn    = normalize(cp)
            c_agg = normalize_name(cp)        # ← normalize_name
            if not cn or not c_agg: continue

            # ── مفتاح التكرار: يبقى داخل نفس المنافس فقط حتى لا نفقد نفس المنتج عند منافسين مختلفين ──────
            bare_ck = re.sub(r"\btester\b|تستر|tester", "", c_agg).strip()
            if not bare_ck or len(bare_ck) < 3:
                continue
            comp_sku_raw = _pid(row, icol)
            comp_sku_key = _norm_sku_barrier(comp_sku_raw)
            comp_url_key = (_cell_str(row, url_col).strip().lower() if url_col else "")
            dedupe_key = (
                str(cname or "").strip().lower(),
                comp_sku_key or bare_ck,
                comp_url_key or bare_ck,
            )
            if dedupe_key in seen_bare:
                continue

            c_brand   = extract_brand(cp)
            c_pline   = extract_product_line(cp, c_brand)
            c_size    = extract_size(cp)
            c_type    = extract_type(cp)
            c_gender  = extract_gender(cp)
            c_is_t    = is_tester(cp)

            # v31.11c: فلاتر الدقة — نستبعد المنتجات غير العطرية والمجموعات
            _cp_class = classify_product(cp)
            _BAD_CLASSES = ('deodorant','hair_mist','body_mist','body_lotion',
                           'soap','shower_gel','after_shave','rejected','other')
            if _cp_class in _BAD_CLASSES:
                continue  # ليس عطر → لا نريده في المفقودات
            # رفض المجموعات/الطقم — لا نريد إضافة مجموعات
            _cp_low = cp.lower()
            if any(w in _cp_low for w in ('مجموعة','مجموعه','طقم','gift set','gift box','set ')):
                continue
            # رفض الأسعار المتطرفة (أقل من 20 ريال أو أكثر من 15000)
            _cp_price = _price(row)
            if _cp_price > 0 and (_cp_price < 20 or _cp_price > 15000):
                continue
            # رفض الأسماء القصيرة جداً (أقل من 8 حروف)
            if len(cp.strip()) < 8:
                continue
            # ── الإصلاح 1: فلتر الحجم — عطر بلا حجم أو ميني < 10مل لا يدخل المفقودة ──
            if not c_size or c_size <= 0:
                _drop_nosize += 1
                continue
            if c_size < _MIN_SIZE_ML:
                _drop_mini += 1
                continue

            # ── Cross-check الأول: بالـ normalize_aggressive ─────────
            found, score, reason, variant = _is_same_product(
                cp, cn, c_brand, c_pline, c_size, c_type, c_gender, c_is_t, c_agg)

            if found:
                continue  # موجود لدينا → تخطي

            # ── Cross-check الثاني: token_set_ratio المباشر على bare ─
            # يحمي من الحالات الهامشية التي يفوتها _is_same_product
            if not found:
                for p in our_items:
                    direct = fuzz.token_set_ratio(bare_ck, p["bare"])
                    if direct >= 82:   # 82% بعد الـ normalize_aggressive = تطابق فعلي
                        found = True
                        break

            if found:
                continue

            seen_bare.add(dedupe_key)

            # ── حساب درجة الثقة ──────────────────────────────
            # score = أعلى نسبة تشابه مع منتجاتنا (كلما انخفضت = مفقود مؤكد أكثر)
            # FIX: إعادة ترتيب الفروع ومنع التصنيف الخاطئ لـ "green" عند score مرتفع.
            # الـ else النهائي كان يحوّل أي منتج ≥68% إلى "green" (مفقود مؤكد)،
            # مما ضخّم عدد المفقودات بشكل خاطئ. الآن نصنّف:
            #   - variant "similar" أو _has_similar (⚠️) → red/yellow
            #   - score ≥68% بلا دلالة مفقود → yellow (ليس مؤكد)
            #   - score < 55 بلا شبيه/variant → green (مؤكد)
            #   - 55 ≤ score < 68 → yellow (محتمل)
            _has_similar = bool(reason and "⚠️" in reason)
            _has_var     = bool(variant)
            _var_type    = (variant or {}).get("type", "")

            if _has_var and _var_type == "similar":
                _conf_level = "red"      # مشكوك — متشابه جداً مع منتج عندنا
            elif _has_similar:
                _conf_level = "yellow"   # ملاحظة تحذير → محتمل
            elif score < 55 and not _has_var:
                _conf_level = "green"    # مفقود مؤكد — فرق واضح عن كل كتالوجنا
            elif score < 68:
                _conf_level = "yellow"   # مفقود محتمل — يحتاج تحقق
            elif _has_var:
                # نوع متاح (تستر/أساسي) موجود عندنا — ليس مفقوداً مؤكداً
                _conf_level = "yellow"
            else:
                # score ≥ 68% بدون variant/similar: تشابه عالٍ
                # = فرصة مفقودة محتملة لكن ليست "مؤكدة"
                _conf_level = "yellow"

            _img_url = _extract_image_url_from_cell(row.get(img_col)) if img_col else ""
            if not _img_url:
                _img_url = _first_image_url_from_row(row)
            _rlink = _cell_str(row, url_col) if url_col else ""
            if not (_rlink and _rlink.startswith("http")):
                _rlink = _first_product_page_url_from_row(row)
            # ملفات بلا عمود صورة (رابط صفحة فقط مثل مهلة): og:image ثم أيقونة الموقع
            if not _img_url and _rlink and _rlink.startswith("http"):
                _try_og = fetch_og_image_url(_rlink)
                if _try_og:
                    _img_url = _try_og
            if not _img_url and _rlink and _rlink.startswith("http"):
                _img_url = favicon_url_for_site(_rlink)
            entry = {
                "منتج_المنافس":  cp,
                "معرف_المنافس":  _pid(row, icol),
                "سعر_المنافس":   _price(row),
                "المنافس":       cname,
                "الماركة":       c_brand,
                "الحجم":         f"{int(c_size)}ml" if c_size else "",
                "النوع":         c_type,
                "الجنس":         c_gender,
                "هو_تستر":       c_is_t,
                "تاريخ_الرصد":   datetime.now().strftime("%Y-%m-%d"),
                "ملاحظة":        reason if reason and "⚠️" in reason else "",
                "درجة_التشابه":  round(score, 1),
                "مستوى_الثقة":  _conf_level,
                "صورة_المنافس":  _img_url,
                "رابط_المنافس":  _rlink,
                "تصنيف_المنتج": classify_product_category(cp),
            }

            # إضافة معلومات النوع المتاح (تستر/أساسي)
            if variant:
                entry["نوع_متاح"]       = variant.get("label","")
                entry["منتج_متاح"]      = variant.get("product","")
                entry["نسبة_التشابه"]   = round(variant.get("score", 0), 1)
            else:
                entry["نوع_متاح"]       = ""
                entry["منتج_متاح"]      = ""
                entry["نسبة_التشابه"]   = 0.0

            missing.append(entry)

    # ── دمج عالمي (Global Dedup) ────────────────────────────────────────
    # نفس المنتج يَرِد مرة لكل منافس (5 منافسين = 5 سجلات). ندمجها في سجل
    # واحد: المفتاح = (الماركة المطبَّعة + الاسم العاري بعد إزالة الحجم/التركيز).
    # نُبقي أقل سعر، ونحفظ أسماء كل المنافسين في حقل «المنافسون».
    _before_dedup = len(missing)
    if missing:
        _merged: dict = {}
        for _e in missing:
            _bn = re.sub(r"\btester\b|تستر|tester", "",
                         normalize_name(_e.get("منتج_المنافس", ""))).strip()
            if not _bn:
                _bn = normalize(_e.get("منتج_المنافس", ""))
            _key = (normalize(_e.get("الماركة", "")), _bn)
            _cn = str(_e.get("المنافس", "") or "").strip()
            _pr_new = float(_e.get("سعر_المنافس", 0) or 0)
            # ═══ v33: بناء كائن المنافس الكامل ═══
            _comp_detail = {
                "المنافس":    _cn,
                "اسم_المنتج": str(_e.get("منتج_المنافس", "")),
                "السعر":      _pr_new,
                "الصورة":     str(_e.get("صورة_المنافس", "") or ""),
                "الرابط":     str(_e.get("رابط_المنافس", "") or ""),
                "الحجم":      str(_e.get("الحجم", "") or ""),
                "النوع":      str(_e.get("النوع", "") or ""),
                "المعرف":     str(_e.get("معرف_المنافس", "") or ""),
            }
            _ex = _merged.get(_key)
            if _ex is None:
                # ── أول ظهور: أنشئ القوائم ──
                _e["المنافسون"] = [_cn] if _cn else []
                _e["تفاصيل_المنافسين"] = [_comp_detail]
                _merged[_key] = _e
                continue
            # ── منتج مكرر من منافس آخر: ألحق البيانات الكاملة ──
            if _cn and _cn not in _ex["المنافسون"]:
                _ex["المنافسون"].append(_cn)
            # إلحاق السجل الكامل (منع التكرار بالاسم)
            _existing_stores = {d.get("المنافس", "") for d in _ex.get("تفاصيل_المنافسين", [])}
            if _cn not in _existing_stores:
                _ex.setdefault("تفاصيل_المنافسين", []).append(_comp_detail)
            # ── إبقاء السجل ذو السعر الأرخص كممثل (توافق خلفي) ──
            _pr_old = float(_ex.get("سعر_المنافس", 0) or 0)
            if _pr_new > 0 and (_pr_old <= 0 or _pr_new < _pr_old):
                _saved_details = _ex.get("تفاصيل_المنافسين", [])
                _saved_names = _ex.get("المنافسون", [])
                _e["تفاصيل_المنافسين"] = _saved_details
                _e["المنافسون"] = _saved_names
                _merged[_key] = _e
        missing = list(_merged.values())
        # ── ترتيب تفاصيل المنافسين تصاعدياً بالسعر + حساب العدد ──
        for _e in missing:
            _details = _e.get("تفاصيل_المنافسين", [])
            if isinstance(_details, list) and len(_details) > 1:
                _details.sort(key=lambda x: float(x.get("السعر", 0) or 999999))
            _e["عدد_المنافسين"] = len(_details) if isinstance(_details, list) else 1
            # ── حساب نطاق الأسعار للعرض السريع ──
            _valid_prices = [float(d.get("السعر", 0) or 0) for d in _details if float(d.get("السعر", 0) or 0) > 0] if isinstance(_details, list) else []
            if _valid_prices:
                _e["أقل_سعر"] = min(_valid_prices)
                _e["أعلى_سعر"] = max(_valid_prices)
                _e["متوسط_السعر"] = round(sum(_valid_prices) / len(_valid_prices), 1)
            else:
                _e["أقل_سعر"] = float(_e.get("سعر_المنافس", 0) or 0)
                _e["أعلى_سعر"] = _e["أقل_سعر"]
                _e["متوسط_السعر"] = _e["أقل_سعر"]
        # ── تحويل قائمة الأسماء إلى نص (التوافق الخلفي) ──
        for _e in missing:
            _lst = _e.get("المنافسون", [])
            _e["المنافسون"] = "، ".join([x for x in _lst if x]) if isinstance(_lst, list) else str(_lst or "")
    _after_dedup = len(missing)
    try:
        logging.getLogger("engines.engine").info(
            f"[find_missing_products] فلتر الحجم: بلا_حجم={_drop_nosize} "
            f"ميني<10مل={_drop_mini} | دمج عالمي: قبل={_before_dedup} "
            f"بعد={_after_dedup} مُدمَج={_before_dedup - _after_dedup}"
        )
    except Exception:
        pass

    return pd.DataFrame(missing) if missing else pd.DataFrame()


def prepare_missing_for_upload(missing_df, margin_pct=15):
    """
    v31.11c: تجهيز المنتجات المفقودة المؤكدة (green) للرفع في المتجر.
    
    يُنشئ DataFrame بأعمدة جاهزة لقالب سلة:
    - اسم المنتج (منظف)
    - الماركة  
    - التصنيف (عطور رجالية/نسائية/للجنسين)
    - الحجم
    - التركيز (EDP/EDT/etc)
    - السعر المقترح (سعر المنافس + هامش)
    - الوصف المنسق
    - صورة المنتج
    - SKU مقترح
    """
    if not isinstance(missing_df, pd.DataFrame) or missing_df.empty:
        return pd.DataFrame()
    
    # فلتر المؤكدين فقط (green)
    if "مستوى_الثقة" in missing_df.columns:
        green = missing_df[missing_df["مستوى_الثقة"] == "green"].copy()
    else:
        green = missing_df.copy()
    
    if green.empty:
        return pd.DataFrame()
    
    rows = []
    _used_skus = set()
    
    for _, r in green.iterrows():
        raw_name = str(r.get("منتج_المنافس", "")).strip()
        if not raw_name:
            continue
        
        brand = str(r.get("الماركة", "") or extract_brand(raw_name) or "")
        size_raw = str(r.get("الحجم", "") or "")
        size_ml = extract_size(raw_name)
        ptype = str(r.get("النوع", "") or extract_type(raw_name) or "")
        gender = str(r.get("الجنس", "") or extract_gender(raw_name) or "")
        comp_price = float(r.get("سعر_المنافس", 0) or 0)
        img_url = str(r.get("صورة_المنافس", "") or "")
        comp_url = str(r.get("رابط_المنافس", "") or "")
        comp_name = str(r.get("المنافس", "") or "")
        
        # ── تنظيف اسم المنتج ──
        clean_name = raw_name
        # إزالة كلمات المنافس الزائدة
        for junk in ["عطر","perfume","عطور"]:
            if clean_name.lower().startswith(junk):
                clean_name = clean_name[len(junk):].strip()
        # إعادة إضافة "عطر" بشكل موحد
        if not clean_name.lower().startswith("عطر"):
            clean_name = f"عطر {clean_name}"
        
        # ── التصنيف ──
        if "رجال" in gender.lower() or "male" in gender.lower() or "homme" in gender.lower():
            category = "عطور رجالية"
            gender_ar = "رجالي"
        elif "نسا" in gender.lower() or "female" in gender.lower() or "femme" in gender.lower():
            category = "عطور نسائية"
            gender_ar = "نسائي"
        else:
            category = "عطور للجنسين"
            gender_ar = "للجنسين"
        
        # ── التركيز بالعربي ──
        type_map = {
            "EDP": "أو دو برفيوم", "EDT": "أو دو تواليت",
            "PARFUM": "برفيوم/اكستريت", "EDC": "أو دو كولون",
            "ELIXIR": "إليكسير", "EXCLUSIF": "إكسكلوسيف",
        }
        type_ar = type_map.get(ptype, ptype or "أو دو برفيوم")
        
        # ── السعر المقترح ──
        if comp_price > 0:
            suggested_price = round(comp_price * (1 + margin_pct / 100), 0)
        else:
            suggested_price = 0
        
        # ── SKU ──
        brand_code = re.sub(r'[^\w]', '', brand[:4]).upper() or "PRF"
        size_code = str(int(size_ml)) if size_ml else "0"
        import hashlib
        _hash = hashlib.md5(raw_name.encode()).hexdigest()[:4].upper()
        sku = f"NEW-{brand_code}-{size_code}-{_hash}"
        while sku in _used_skus:
            sku += "X"
        _used_skus.add(sku)
        
        # ── الوصف المنسق ──
        desc_parts = [f"عطر {brand}" if brand else "عطر"]
        if size_ml:
            desc_parts.append(f"الحجم: {int(size_ml)} مل")
        if type_ar:
            desc_parts.append(f"التركيز: {type_ar}")
        if gender_ar:
            desc_parts.append(f"مناسب: {gender_ar}")
        description = " | ".join(desc_parts)
        
        rows.append({
            "اسم_المنتج": clean_name,
            "الماركة": brand,
            "التصنيف": category,
            "الحجم_مل": int(size_ml) if size_ml else "",
            "التركيز": type_ar,
            "الجنس": gender_ar,
            "سعر_المنافس": comp_price,
            "السعر_المقترح": suggested_price,
            "الوصف": description,
            "صورة_المنتج": img_url,
            "رابط_المصدر": comp_url,
            "المنافس_المصدر": comp_name,
            "SKU": sku,
            "الحالة": "جاهز_للرفع",
            "تاريخ_الرصد": str(r.get("تاريخ_الرصد", "")),
        })
    
    return pd.DataFrame(rows) if rows else pd.DataFrame()


# ═══════════════════════════════════════════════════════
#  محرك أولويات المفقودات الذكي v33
# ═══════════════════════════════════════════════════════
def calculate_missing_priority(missing_df: pd.DataFrame) -> pd.DataFrame:
    """
    يحسب درجة أولوية (0-100) لكل منتج مفقود.
    الأولوية = (عدد_المنافسين × 40%) + (مستوى_الثقة × 25%)
             + (جاذبية_السعر × 20%) + (اكتمال_البيانات × 15%)
    """
    if not isinstance(missing_df, pd.DataFrame) or missing_df.empty:
        return missing_df
    df = missing_df.copy()
    # ═══ 1. عدد المنافسين (40%) ═══
    def _count(x):
        if isinstance(x, list): return len(x)
        if isinstance(x, str) and x.strip().startswith("["):
            try: return len(__import__("json").loads(x))
            except Exception: return 1
        return 1
    if "تفاصيل_المنافسين" in df.columns:
        _cc = df["تفاصيل_المنافسين"].apply(_count)
    elif "عدد_المنافسين" in df.columns:
        _cc = pd.to_numeric(df["عدد_المنافسين"], errors="coerce").fillna(1)
    elif "المنافسون" in df.columns:
        _cc = df["المنافسون"].apply(lambda x: len(str(x).split("،")) if x and str(x).strip() else 1)
    else:
        _cc = pd.Series(1, index=df.index)
    _max_c = max(_cc.max(), 1)
    s_comp = (_cc / _max_c) * 40
    # ═══ 2. مستوى الثقة (25%) ═══
    _cmap = {"green": 25, "review": 15, "yellow": 8, "red": 2}
    if "مستوى_الثقة" in df.columns:
        s_conf = df["مستوى_الثقة"].map(_cmap).fillna(8)
    else:
        s_conf = pd.Series(25, index=df.index)  # افتراضي: مؤكد
    # ═══ 3. جاذبية السعر (20%) ═══
    _pc = None
    for c in ("سعر_المنافس", "أقل_سعر", "السعر"):
        if c in df.columns:
            _pc = c
            break
    if _pc:
        _prices = pd.to_numeric(df[_pc], errors="coerce").fillna(0)
        s_price = _prices.apply(lambda p: (
            20 if 60 <= p <= 500 else
            14 if 30 <= p <= 800 else
            8 if p > 0 else 0
        ))
    else:
        s_price = pd.Series(10, index=df.index)
    # ═══ 4. اكتمال البيانات (15%) ═══
    def _data_score(row):
        s = 0
        if str(row.get("الماركة", "") or "").strip(): s += 4
        if str(row.get("الحجم", "") or "").strip(): s += 3
        if str(row.get("صورة_المنافس", "") or "").strip().startswith("http"): s += 4
        if str(row.get("النوع", "") or "").strip(): s += 2
        if str(row.get("الجنس", "") or "").strip(): s += 2
        return min(s, 15)
    s_data = df.apply(_data_score, axis=1)
    # ═══ المجموع ═══
    df["درجة_الأولوية"] = (s_comp + s_conf + s_price + s_data).round(0).astype(int).clip(0, 100)
    df["مستوى_الأولوية"] = df["درجة_الأولوية"].apply(lambda s: (
        "🔴 حرج"     if s >= 80 else
        "🟠 عالي"    if s >= 60 else
        "🟡 متوسط"   if s >= 40 else
        "🟢 منخفض"
    ))
    return df


def export_excel(df, sheet_name="النتائج"):
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    output = io.BytesIO()
    edf = df.copy()
    for col in ["جميع المنافسين","جميع_المنافسين"]:
        if col in edf.columns: edf = edf.drop(columns=[col])
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        edf.to_excel(writer, sheet_name=sheet_name[:31], index=False)
        ws = writer.sheets[sheet_name[:31]]
        hfill = PatternFill("solid", fgColor="1a1a2e")
        hfont = Font(color="FFFFFF", bold=True, size=10)
        for cell in ws[1]:
            cell.fill=hfill; cell.font=hfont
            cell.alignment=Alignment(horizontal="center")
        # تم تعديل المسميات هنا لمطابقة طلبك بدقة تامة
        COLORS = {"🔴 سعر أعلى":"FFCCCC","🟢 سعر أقل":"CCFFCC",
                  "✅ موافق":"CCFFEE","⚠️ تحت المراجعة":"FFF3CC","🔍 منتجات مفقودة":"CCE5FF"}
        dcol = None
        for i, cell in enumerate(ws[1], 1):
            if cell.value and "القرار" in str(cell.value): dcol=i; break
        if dcol:
            for ri, row in enumerate(ws.iter_rows(min_row=2), 2):
                val = str(ws.cell(ri,dcol).value or "")
                for k,c in COLORS.items():
                    if k.split()[0] in val:
                        for cell in row: cell.fill=PatternFill("solid",fgColor=c)
                        break
        for ci, col in enumerate(ws.columns, 1):
            w = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(ci)].width = min(w+4, 55)
    return output.getvalue()

def export_section_excel(df, sname):
    return export_excel(df, sheet_name=sname[:31])


# ═══════════════════════════════════════════════════════
#  الحاجز الذكي للمفقودات (Double-Barrier) — مطابقة SKU + Fuzzy
# ═══════════════════════════════════════════════════════
def _norm_sku_barrier(s) -> str:
    if s is None:
        return ""
    t = str(s).strip()
    if not t or t.lower() in ("nan", "none", "0", "0.0"):
        return ""
    try:
        return str(int(float(t)))
    except (ValueError, TypeError):
        return t


def _our_product_names_series(our_df: pd.DataFrame):
    c = _name_col_for_analysis(our_df)
    if c and c in our_df.columns:
        return our_df[c].dropna().astype(str).tolist()
    return []


def _our_sku_set(our_df: pd.DataFrame) -> set:
    out = set()
    for c in [
        "رقم_المنتج",
        "رقم المنتج",
        "معرف_المنتج",
        "معرف المنتج",
        "SKU",
        "sku",
    ]:
        if c not in our_df.columns:
            continue
        for v in our_df[c].dropna().astype(str):
            ns = _norm_sku_barrier(v)
            if ns:
                out.add(ns)
                out.add(str(v).strip())
    return out


def smart_missing_barrier(missing_df: pd.DataFrame, our_df: pd.DataFrame, threshold: int = 92) -> pd.DataFrame:
    """
    محرك الحاجز الذكي: الفلتر النهائي قبل دخول المنتجات لقسم المفقودات.
    يضمن عدم تكرار عبر مطابقة الـ SKU والـ Fuzzy Matching الصارم مع كتالوجنا.
    """
    if missing_df.empty:
        return missing_df

    filtered_df, _ = apply_strict_pipeline_filters(missing_df, name_col="منتج_المنافس")

    if filtered_df.empty:
        return filtered_df

    if our_df is None or our_df.empty:
        return filtered_df.reset_index(drop=True)

    our_names = _our_product_names_series(our_df)
    if not our_names:
        return filtered_df.reset_index(drop=True)

    our_skus = _our_sku_set(our_df)

    keep_idx = []
    for idx, row in filtered_df.iterrows():
        comp_sku = _norm_sku_barrier(row.get("معرف_المنافس", ""))
        raw_sku = str(row.get("معرف_المنافس", "")).strip()
        comp_name = str(row.get("منتج_المنافس", "")).strip()

        if comp_sku and (comp_sku in our_skus or raw_sku in our_skus):
            continue

        match = rf_process.extractOne(comp_name, our_names, scorer=fuzz.token_set_ratio)
        if match and match[1] >= threshold:
            continue

        keep_idx.append(idx)

    if not keep_idx:
        return pd.DataFrame()

    return filtered_df.loc[keep_idx].reset_index(drop=True)
